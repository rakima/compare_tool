"""Create sample files for manually checking compare_tool."""

import csv
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

ROOT = Path(__file__).resolve().parents[1]
OUTPUT_DIR = ROOT / "samples"
EXCEL_97_2003_FILE_FORMAT = 56


def add_common_sheet(workbook: Workbook, *, new: bool) -> None:
    sheet = workbook.active
    sheet.title = "商品一覧"
    sheet.append(["商品ID", "商品名", "単価", "数量", "金額", "備考"])
    rows = (
        (
            ["P001", "ノートPC", 120000 if new else 100000, 2, "=C2*D2", "価格変更" if new else ""],
            ["P002", "マウス", 2500, 10 if new else 8, "=C3*D3", "数量変更"],
            ["P004", "Webカメラ", 8500, 4, "=C4*D4", "追加セル・行"],
        )
        if new
        else (
            ["P001", "ノートPC", 100000, 2, "=C2*D2", ""],
            ["P002", "マウス", 2500, 8, "=C3*D3", "数量変更"],
            ["P003", "キーボード", 6000, 5, "=C4*D4", "新ファイルでは削除"],
        )
    )
    for row in rows:
        sheet.append(row)

    # A formula-only change that keeps the displayed intent obvious.
    sheet["H1"] = "数式変更確認"
    sheet["H2"] = "=SUM(E2:E4)" if new else "=SUM(E2:E3)"
    if new:
        sheet["J10"] = "新ファイルで追加されたセル"
    else:
        sheet["J11"] = "新ファイルでは削除されたセル"

    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
    sheet.freeze_panes = "A2"
    for column, width in {"A": 12, "B": 18, "C": 12, "D": 10, "E": 14, "F": 22, "H": 18, "J": 28}.items():
        sheet.column_dimensions[column].width = width


def add_option_sheet(workbook: Workbook, *, new: bool) -> None:
    sheet = workbook.create_sheet("オプション確認")
    sheet.append(["確認項目", "値", "期待結果"])
    sheet.append(["前後スペース", " Sample " if not new else "Sample", "無視ONなら差分なし"])
    sheet.append(["大文字小文字", "EXCEL" if not new else "excel", "無視ONなら差分なし"])
    sheet.append(["空セル", "" if not new else None, "同一視ONなら差分なし"])
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
    sheet.column_dimensions["A"].width = 20
    sheet.column_dimensions["B"].width = 18
    sheet.column_dimensions["C"].width = 24


def create_workbook(path: Path, *, new: bool) -> None:
    workbook = Workbook()
    add_common_sheet(workbook, new=new)
    add_option_sheet(workbook, new=new)
    unique_name = "新規シート" if new else "廃止シート"
    unique = workbook.create_sheet(unique_name)
    unique["A1"] = "新ファイルで追加されたシート" if new else "新ファイルでは削除されるシート"
    workbook.save(path)
    workbook.close()


def create_xls_from_xlsx(source: Path, output: Path) -> bool:
    """Create an Excel 97-2003 workbook using Excel COM when available."""
    try:
        import win32com.client
    except ImportError:
        print("Skipped .xls samples: pywin32 is not installed.")
        return False

    excel = None
    workbook = None
    try:
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        workbook = excel.Workbooks.Open(str(source.resolve()), UpdateLinks=0, ReadOnly=True)
        workbook.SaveAs(str(output.resolve()), FileFormat=EXCEL_97_2003_FILE_FORMAT)
    except Exception as exc:
        print(f"Skipped .xls sample {output.name}: Excel conversion failed: {exc}")
        return False
    finally:
        if workbook is not None:
            workbook.Close(False)
        if excel is not None:
            excel.Quit()
    return True


def create_csv(path: Path, *, new: bool, encoding: str = "utf-8-sig", delimiter: str = ",") -> None:
    rows = (
        [
            ["商品ID", "商品名", "単価", "数量", "備考"],
            ["P001", "ノートPC", "120000" if new else "100000", "2", "価格変更" if new else ""],
            ["P002", "マウス", "2500", "10" if new else "8", "数量変更"],
            ["P004", "Webカメラ", "8500", "4", "追加行"],
        ]
        if new
        else [
            ["商品ID", "商品名", "単価", "数量", "備考"],
            ["P001", "ノートPC", "100000", "2", ""],
            ["P002", "マウス", "2500", "8", "数量変更"],
            ["P003", "キーボード", "6000", "5", "削除行"],
        ]
    )
    with path.open("w", encoding=encoding, newline="") as stream:
        writer = csv.writer(stream, delimiter=delimiter)
        writer.writerows(rows)


def create_json(path: Path, *, new: bool) -> None:
    data = (
        {
            "product": {
                "id": "P001",
                "name": "ノートPC",
                "price": 120000,
                "tags": ["pc", "business", "new"],
            },
            "settings": {
                "enabled": True,
                "mode": "standard",
            },
            "items": [
                {"id": "P001", "quantity": 2},
                {"id": "P002", "quantity": 10},
                {"id": "P004", "quantity": 4},
            ],
            "added": "新JSONで追加されたキー",
        }
        if new
        else {
            "product": {
                "id": "P001",
                "name": "ノートPC",
                "price": 100000,
                "tags": ["pc", "business"],
            },
            "settings": {
                "enabled": True,
                "mode": "standard",
            },
            "items": [
                {"id": "P001", "quantity": 2},
                {"id": "P002", "quantity": 8},
            ],
            "removed": "新JSONでは削除されたキー",
        }
    )
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def create_xml(path: Path, *, new: bool) -> None:
    text = (
        """<?xml version="1.0" encoding="UTF-8"?>
<catalog enabled="false" added="yes">
  <product id="P001">
    <name>ノートPC</name>
    <price>120000</price>
    <tags>
      <tag>pc</tag>
      <tag>business</tag>
      <tag>new</tag>
    </tags>
  </product>
  <items>
    <item id="P001" quantity="2" />
    <item id="P002" quantity="10" />
    <item id="P004" quantity="4" />
  </items>
  <suppliers>
    <supplier code="S002" rank="A">西日本商事</supplier>
    <supplier code="S001" rank="A">東日本商事</supplier>
  </suppliers>
  <note>Sample</note>
</catalog>
"""
        if new
        else """<?xml version="1.0" encoding="UTF-8"?>
<catalog enabled="true" removed="yes">
  <product id="P001">
    <name>ノートPC</name>
    <price>100000</price>
    <tags>
      <tag>pc</tag>
      <tag>business</tag>
    </tags>
  </product>
  <items>
    <item id="P001" quantity="2" />
    <item id="P002" quantity="8" />
    <item id="P003" quantity="5" />
  </items>
  <suppliers>
    <supplier code="S001" rank="A">東日本商事</supplier>
    <supplier code="S002" rank="B">西日本商事</supplier>
  </suppliers>
  <note> Sample </note>
</catalog>
"""
    )
    path.write_text(text, encoding="utf-8")


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    old_xlsx = OUTPUT_DIR / "比較サンプル_旧.xlsx"
    new_xlsx = OUTPUT_DIR / "比較サンプル_新.xlsx"
    create_workbook(old_xlsx, new=False)
    create_workbook(new_xlsx, new=True)
    print("Created .xlsx samples.")
    if create_xls_from_xlsx(old_xlsx, OUTPUT_DIR / "比較サンプル_旧.xls") and create_xls_from_xlsx(
        new_xlsx,
        OUTPUT_DIR / "比較サンプル_新.xls",
    ):
        print("Created .xls samples.")
    create_csv(OUTPUT_DIR / "比較サンプル_旧.csv", new=False)
    create_csv(OUTPUT_DIR / "比較サンプル_新.csv", new=True)
    create_csv(OUTPUT_DIR / "比較サンプル_ShiftJIS_Tab_旧.csv", new=False, encoding="cp932", delimiter="\t")
    create_csv(OUTPUT_DIR / "比較サンプル_ShiftJIS_Tab_新.csv", new=True, encoding="cp932", delimiter="\t")
    print("Created .csv samples.")
    create_json(OUTPUT_DIR / "比較サンプル_旧.json", new=False)
    create_json(OUTPUT_DIR / "比較サンプル_新.json", new=True)
    print("Created .json samples.")
    create_xml(OUTPUT_DIR / "比較サンプル_旧.xml", new=False)
    create_xml(OUTPUT_DIR / "比較サンプル_新.xml", new=True)
    print("Created .xml samples.")
    print(f"Created sample files in {OUTPUT_DIR}")


if __name__ == "__main__":
    main()
