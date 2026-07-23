"""End-to-end acceptance tests using real .xlsx files."""

import json
from collections.abc import Mapping
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

from compare_tool.errors import (
    InvalidInputError,
    OperationCancelledError,
    OutputWriteError,
    WorkbookConversionError,
    WorkbookReadError,
)
from compare_tool.excel import ExcelReportWriter
from compare_tool.models import CompareAlgorithm, CompareOptions, CompareResult, Difference, DifferenceType
from compare_tool.settings import AppSettingsStore
from compare_tool.usecase import CompareUseCase
from compare_tool.workbook_preparer import ExcelWorkbookConverter, PreparedWorkbook, WorkbookPreparer


def make_workbook(path: Path, sheets: Mapping[str, Mapping[str, object]]) -> Path:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for name, cells in sheets.items():
        sheet = workbook.create_sheet(name)
        for coordinate, value in cells.items():
            sheet[coordinate] = value
    workbook.save(path)
    workbook.close()
    return path


def make_csv(
    path: Path,
    rows: list[list[object]],
    *,
    encoding: str = "utf-8",
    delimiter: str = ",",
) -> Path:
    path.write_text(
        "\n".join(delimiter.join(str(value) for value in row) for row in rows),
        encoding=encoding,
    )
    return path


def make_json(path: Path, data: object) -> Path:
    path.write_text(json.dumps(data, ensure_ascii=False), encoding="utf-8")
    return path


def make_xml(path: Path, text: str) -> Path:
    path.write_text(text, encoding="utf-8")
    return path


def compare(
    tmp_path: Path,
    old_sheets: Mapping[str, Mapping[str, object]],
    new_sheets: Mapping[str, Mapping[str, object]],
    *,
    detailed: bool = True,
    options: CompareOptions | None = None,
) -> tuple[CompareResult, Path]:
    old = make_workbook(tmp_path / "old.xlsx", old_sheets)
    new = make_workbook(tmp_path / "new.xlsx", new_sheets)
    output = tmp_path / "output.xlsx"
    result = CompareUseCase().execute(old, new, output, options or CompareOptions(), detailed)
    return result, output


def test_same_input_file_is_rejected(tmp_path: Path) -> None:
    source = make_workbook(tmp_path / "same.xlsx", {"Data": {"A1": "value"}})
    with pytest.raises(InvalidInputError, match="同じファイル"):
        CompareUseCase().execute(source, source, tmp_path / "output.xlsx", CompareOptions())


@pytest.mark.parametrize("bad_name", ["input.xlsm", "input.yaml"])
def test_unsupported_input_extension_is_rejected(tmp_path: Path, bad_name: str) -> None:
    bad = tmp_path / bad_name
    bad.write_bytes(b"not used")
    new = make_workbook(tmp_path / "new.xlsx", {"Data": {}})
    with pytest.raises(InvalidInputError, match=r"\.xlsx、\.xls、\.csv、\.json、\.xml"):
        CompareUseCase().execute(bad, new, tmp_path / "output.xlsx", CompareOptions())


def test_mixed_input_formats_are_rejected(tmp_path: Path) -> None:
    old = make_csv(tmp_path / "old.csv", [["id", "value"], [1, "old"]])
    new = make_workbook(tmp_path / "new.xlsx", {"Data": {"A1": "id"}})

    with pytest.raises(InvalidInputError, match="同じ形式"):
        CompareUseCase().execute(old, new, tmp_path / "output.xlsx", CompareOptions())


def test_xls_conversion_failure_is_reported(tmp_path: Path) -> None:
    class FailingConverter(ExcelWorkbookConverter):
        def convert(self, path: Path) -> Path:
            raise WorkbookConversionError("simulated conversion failure")

    old = tmp_path / "old.xls"
    old.write_bytes(b"legacy workbook placeholder")
    new = make_workbook(tmp_path / "new.xlsx", {"Data": {}})

    with pytest.raises(WorkbookConversionError, match="simulated conversion failure"):
        CompareUseCase(workbook_preparer=WorkbookPreparer(FailingConverter())).execute(
            old,
            new,
            tmp_path / "output.xlsx",
            CompareOptions(),
        )


def test_xls_input_can_be_prepared_by_injected_converter(tmp_path: Path) -> None:
    class FakeWorkbookPreparer(WorkbookPreparer):
        def __init__(self, converted_old: Path) -> None:
            self.converted_old = converted_old

        def prepare(self, path: Path) -> PreparedWorkbook:
            if path.suffix.lower() == ".xls":
                return PreparedWorkbook(path, self.converted_old, converted=True)
            return super().prepare(path)

    old_xls = tmp_path / "old.xls"
    old_xls.write_bytes(b"legacy workbook placeholder")
    converted_old = make_workbook(tmp_path / "converted_old.xlsx", {"Data": {"A1": "old"}})
    new = make_workbook(tmp_path / "new.xlsx", {"Data": {"A1": "new"}})
    output = tmp_path / "output.xlsx"

    result = CompareUseCase(workbook_preparer=FakeWorkbookPreparer(converted_old)).execute(
        old_xls,
        new,
        output,
        CompareOptions(),
    )

    assert result.count(DifferenceType.MODIFIED) == 1
    workbook = load_workbook(output)
    assert workbook["比較結果"]["D11"].value == "old"
    assert workbook["比較結果"]["E11"].value == "new"
    workbook.close()


def test_xls_inputs_can_both_be_prepared_by_injected_converter(tmp_path: Path) -> None:
    class FakeWorkbookPreparer(WorkbookPreparer):
        def __init__(self, converted_old: Path, converted_new: Path) -> None:
            self.converted_old = converted_old
            self.converted_new = converted_new

        def prepare(self, path: Path) -> PreparedWorkbook:
            if path.name == "old.xls":
                return PreparedWorkbook(path, self.converted_old, converted=True)
            if path.name == "new.xls":
                return PreparedWorkbook(path, self.converted_new, converted=True)
            return super().prepare(path)

    old_xls = tmp_path / "old.xls"
    new_xls = tmp_path / "new.xls"
    old_xls.write_bytes(b"legacy old workbook placeholder")
    new_xls.write_bytes(b"legacy new workbook placeholder")
    converted_old = make_workbook(tmp_path / "converted_old.xlsx", {"Data": {"A1": "old"}})
    converted_new = make_workbook(tmp_path / "converted_new.xlsx", {"Data": {"A1": "new"}})
    output = tmp_path / "output.xlsx"

    result = CompareUseCase(
        workbook_preparer=FakeWorkbookPreparer(converted_old, converted_new),
    ).execute(old_xls, new_xls, output, CompareOptions())

    assert result.count(DifferenceType.MODIFIED) == 1
    workbook = load_workbook(output)
    assert workbook["Data"]["A1"].value == "new"
    assert workbook["比較結果"]["D11"].value == "old"
    assert workbook["比較結果"]["E11"].value == "new"
    workbook.close()


def test_converted_xls_temporary_files_are_removed_after_success(tmp_path: Path) -> None:
    class TemporaryWorkbookPreparer(WorkbookPreparer):
        def __init__(self) -> None:
            self.converted_old = tmp_path / "old_converted.xlsx"
            self.converted_new = tmp_path / "new_converted.xlsx"

        def prepare(self, path: Path) -> PreparedWorkbook:
            if path.name == "old.xls":
                make_workbook(self.converted_old, {"Data": {"A1": "old"}})
                return PreparedWorkbook(path, self.converted_old, converted=True)
            if path.name == "new.xls":
                make_workbook(self.converted_new, {"Data": {"A1": "new"}})
                return PreparedWorkbook(path, self.converted_new, converted=True)
            return super().prepare(path)

    old_xls = tmp_path / "old.xls"
    new_xls = tmp_path / "new.xls"
    old_xls.write_bytes(b"legacy old workbook placeholder")
    new_xls.write_bytes(b"legacy new workbook placeholder")
    preparer = TemporaryWorkbookPreparer()

    CompareUseCase(workbook_preparer=preparer).execute(
        old_xls,
        new_xls,
        tmp_path / "output.xlsx",
        CompareOptions(),
    )

    assert not preparer.converted_old.exists()
    assert not preparer.converted_new.exists()


def test_converted_xls_temporary_file_is_removed_when_second_prepare_fails(tmp_path: Path) -> None:
    class PartiallyFailingWorkbookPreparer(WorkbookPreparer):
        def __init__(self) -> None:
            self.converted_old = tmp_path / "old_converted.xlsx"

        def prepare(self, path: Path) -> PreparedWorkbook:
            if path.name == "old.xls":
                make_workbook(self.converted_old, {"Data": {"A1": "old"}})
                return PreparedWorkbook(path, self.converted_old, converted=True)
            if path.name == "new.xls":
                raise WorkbookConversionError("new workbook conversion failed")
            return super().prepare(path)

    old_xls = tmp_path / "old.xls"
    new_xls = tmp_path / "new.xls"
    old_xls.write_bytes(b"legacy old workbook placeholder")
    new_xls.write_bytes(b"legacy new workbook placeholder")
    preparer = PartiallyFailingWorkbookPreparer()

    with pytest.raises(WorkbookConversionError, match="new workbook conversion failed"):
        CompareUseCase(workbook_preparer=preparer).execute(
            old_xls,
            new_xls,
            tmp_path / "output.xlsx",
            CompareOptions(),
        )

    assert not preparer.converted_old.exists()


def test_csv_files_are_compared_and_written_to_excel_report(tmp_path: Path) -> None:
    old = make_csv(tmp_path / "old.csv", [["id", "value"], [1, "old"], [2, "keep"]])
    new = make_csv(tmp_path / "new.csv", [["id", "value"], [1, "new"], [2, "keep"], [3, "added"]])
    output = tmp_path / "output.xlsx"

    result = CompareUseCase().execute(old, new, output, CompareOptions())

    assert result.count(DifferenceType.MODIFIED) == 1
    assert result.count(DifferenceType.ADDED) == 2
    workbook = load_workbook(output)
    assert workbook.sheetnames == ["比較結果", "CSV"]
    assert workbook["CSV"]["B2"].value == "new"
    assert workbook["CSV"]["A4"].value == "3"
    report = workbook["比較結果"]
    rows = {report.cell(row, 3).value: row for row in range(11, 14)}
    modified_row = rows["B2"]
    assert report.cell(modified_row, 1).value == "変更"
    assert report.cell(modified_row, 2).value == "CSV"
    assert report.cell(modified_row, 6).hyperlink.target == "#'CSV'!B2"
    workbook.close()


def test_csv_key_column_compare_matches_moved_rows(tmp_path: Path) -> None:
    old = make_csv(tmp_path / "old.csv", [["id", "value"], [1, "old"]])
    new = make_csv(tmp_path / "new.csv", [["id", "value"], [2, "other"], [1, "new"]])

    result = CompareUseCase().execute(
        old,
        new,
        tmp_path / "output.xlsx",
        CompareOptions(algorithm=CompareAlgorithm.KEY_COLUMNS, key_columns=("A",)),
    )

    assert result.count(DifferenceType.MODIFIED) == 1
    assert result.count(DifferenceType.ROW_ADDED) == 1
    modified = [difference for difference in result.differences if difference.kind is DifferenceType.MODIFIED][0]
    assert modified.cell == "B3"
    assert modified.old_value == "old"
    assert modified.new_value == "new"


def test_csv_auto_encoding_detects_shift_jis(tmp_path: Path) -> None:
    old = make_csv(
        tmp_path / "old.csv",
        [["id", "名前"], [1, "東京"]],
        encoding="cp932",
    )
    new = make_csv(
        tmp_path / "new.csv",
        [["id", "名前"], [1, "大阪"]],
        encoding="cp932",
    )
    output = tmp_path / "output.xlsx"

    result = CompareUseCase().execute(old, new, output, CompareOptions())

    assert result.count(DifferenceType.MODIFIED) == 1
    workbook = load_workbook(output)
    assert workbook["CSV"]["B2"].value == "大阪"
    report = workbook["比較結果"]
    assert report["D11"].value == "東京"
    assert report["E11"].value == "大阪"
    workbook.close()


def test_csv_encoding_and_delimiter_options_are_used(tmp_path: Path) -> None:
    old = make_csv(
        tmp_path / "old.csv",
        [["id", "名前"], [1, "東京"]],
        encoding="cp932",
        delimiter="\t",
    )
    new = make_csv(
        tmp_path / "new.csv",
        [["id", "名前"], [1, "大阪"]],
        encoding="cp932",
        delimiter="\t",
    )
    output = tmp_path / "output.xlsx"

    result = CompareUseCase().execute(
        old,
        new,
        output,
        CompareOptions(csv_encoding="cp932", csv_delimiter="\t"),
    )

    assert result.count(DifferenceType.MODIFIED) == 1
    workbook = load_workbook(output)
    assert workbook["CSV"]["B2"].value == "大阪"
    report = workbook["比較結果"]
    assert report["C11"].value == "B2"
    assert report["D11"].value == "東京"
    assert report["E11"].value == "大阪"
    assert report["H1"].value == "CSV読み込み設定"
    assert report["H2"].value == "文字コード"
    assert report["I2"].value == "Shift_JIS"
    assert report["H3"].value == "区切り文字"
    assert report["I3"].value == "タブ"
    assert report["H4"].value == "空行を無視"
    assert report["I4"].value == "はい"
    workbook.close()


def test_csv_auto_delimiter_detects_tab_and_semicolon(tmp_path: Path) -> None:
    tab_old = make_csv(tmp_path / "tab_old.csv", [["id", "value"], [1, "old"]], delimiter="\t")
    tab_new = make_csv(tmp_path / "tab_new.csv", [["id", "value"], [1, "new"]], delimiter="\t")
    semicolon_old = make_csv(tmp_path / "semicolon_old.csv", [["id", "value"], [1, "old"]], delimiter=";")
    semicolon_new = make_csv(tmp_path / "semicolon_new.csv", [["id", "value"], [1, "new"]], delimiter=";")

    tab_result = CompareUseCase().execute(tab_old, tab_new, tmp_path / "tab_output.xlsx", CompareOptions())
    semicolon_result = CompareUseCase().execute(
        semicolon_old,
        semicolon_new,
        tmp_path / "semicolon_output.xlsx",
        CompareOptions(),
    )

    assert tab_result.count(DifferenceType.MODIFIED) == 1
    assert semicolon_result.count(DifferenceType.MODIFIED) == 1


def test_csv_blank_lines_are_ignored_by_default(tmp_path: Path) -> None:
    old = make_csv(tmp_path / "old.csv", [["id", "value"], [], [1, "same"]])
    new = make_csv(tmp_path / "new.csv", [["id", "value"], [1, "same"], []])

    result = CompareUseCase().execute(old, new, tmp_path / "output.xlsx", CompareOptions())

    assert result.total == 0


def test_csv_blank_lines_can_be_compared(tmp_path: Path) -> None:
    old = make_csv(tmp_path / "old.csv", [["id", "value"], [1, "same"]])
    new = make_csv(tmp_path / "new.csv", [["id", "value"], [], [1, "same"]])

    result = CompareUseCase().execute(
        old,
        new,
        tmp_path / "output.xlsx",
        CompareOptions(ignore_csv_blank_lines=False),
    )

    assert result.total > 0


def test_invalid_csv_delimiter_is_rejected(tmp_path: Path) -> None:
    old = make_csv(tmp_path / "old.csv", [["id", "value"]])
    new = make_csv(tmp_path / "new.csv", [["id", "value"]])

    with pytest.raises(InvalidInputError, match="CSV区切り文字"):
        CompareUseCase().execute(
            old,
            new,
            tmp_path / "output.xlsx",
            CompareOptions(csv_delimiter="||"),
        )


def test_csv_auto_encoding_failure_explains_next_action(tmp_path: Path) -> None:
    old = tmp_path / "old.csv"
    new = tmp_path / "new.csv"
    old.write_bytes(b"\xff\xfe\x00\x80")
    new.write_bytes(b"id,value\n1,new\n")

    with pytest.raises(WorkbookReadError, match="UTF-8またはShift_JISで保存し直してください"):
        CompareUseCase().execute(old, new, tmp_path / "output.xlsx", CompareOptions())


def test_json_files_are_compared_and_written_to_excel_report(tmp_path: Path) -> None:
    old = make_json(tmp_path / "old.json", {"name": "old", "removed": True})
    new = make_json(tmp_path / "new.json", {"name": "new", "items": [1, 2]})
    output = tmp_path / "output.xlsx"

    result = CompareUseCase().execute(old, new, output, CompareOptions())

    assert result.count(DifferenceType.MODIFIED) == 1
    assert result.count(DifferenceType.ADDED) == 1
    assert result.count(DifferenceType.DELETED) == 1
    workbook = load_workbook(output)
    assert workbook.sheetnames == ["比較結果", "JSON"]
    report = workbook["比較結果"]
    rows = {report.cell(row, 3).value: row for row in range(11, 14)}
    assert report.cell(rows["$.name"], 1).value == "変更"
    assert report.cell(rows["$.name"], 2).value == "JSON"
    assert report.cell(rows["$.name"], 4).value == "old"
    assert report.cell(rows["$.name"], 5).value == "new"
    assert report.cell(rows["$.name"], 6).value is None
    assert report["H1"].value == "JSON読み込み設定"
    assert workbook["JSON"]["A1"].value == "新JSON"
    workbook.close()


def test_json_object_key_order_option_is_used_and_recorded(tmp_path: Path) -> None:
    old = make_json(tmp_path / "old.json", {"a": 1, "b": 2})
    new = make_json(tmp_path / "new.json", {"b": 2, "a": 1})
    output = tmp_path / "output.xlsx"

    result = CompareUseCase().execute(
        old,
        new,
        output,
        CompareOptions(ignore_json_object_key_order=False),
    )

    assert result.count(DifferenceType.MODIFIED) == 1
    difference = result.differences[0]
    assert difference.cell == "$"
    assert difference.old_value == ["a", "b"]
    assert difference.new_value == ["b", "a"]
    workbook = load_workbook(output)
    report = workbook["比較結果"]
    assert report["C11"].value == "$"
    assert report["D11"].value == '["a", "b"]'
    assert report["E11"].value == '["b", "a"]'
    assert report["H4"].value == "オブジェクトのキー順を無視"
    assert report["I4"].value == "いいえ"
    assert report["H5"].value == "配列順序を無視"
    assert report["I5"].value == "いいえ"
    workbook.close()


def test_json_array_order_can_be_ignored_through_usecase(tmp_path: Path) -> None:
    old = make_json(tmp_path / "old.json", {"items": [{"id": 1}, {"id": 2}]})
    new = make_json(tmp_path / "new.json", {"items": [{"id": 2}, {"id": 1}]})
    output = tmp_path / "output.xlsx"

    result = CompareUseCase().execute(
        old,
        new,
        output,
        CompareOptions(ignore_json_array_order=True),
    )

    assert result.total == 0
    workbook = load_workbook(output)
    report = workbook["比較結果"]
    assert report["A11"].value is None
    assert report["I5"].value == "はい"
    workbook.close()


def test_json_array_order_ignore_still_reports_different_array_contents(tmp_path: Path) -> None:
    old = make_json(tmp_path / "old.json", {"items": [{"id": 1}, {"id": 2}]})
    new = make_json(tmp_path / "new.json", {"items": [{"id": 2}, {"id": 3}]})

    result = CompareUseCase().execute(
        old,
        new,
        tmp_path / "output.xlsx",
        CompareOptions(ignore_json_array_order=True),
    )

    assert result.count(DifferenceType.DELETED) == 1
    assert result.count(DifferenceType.ADDED) == 1
    differences = [
        (difference.kind, difference.cell, difference.old_value, difference.new_value)
        for difference in result.differences
    ]
    assert differences == [
        (DifferenceType.DELETED, "$.items[0]", {"id": 1}, None),
        (DifferenceType.ADDED, "$.items[1]", None, {"id": 3}),
    ]


def test_json_array_key_compare_matches_reordered_objects_through_usecase(tmp_path: Path) -> None:
    old = make_json(
        tmp_path / "old.json",
        {"items": [{"id": "P001", "price": 100}, {"id": "P002", "price": 200}]},
    )
    new = make_json(
        tmp_path / "new.json",
        {"items": [{"id": "P002", "price": 250}, {"id": "P001", "price": 100}]},
    )
    output = tmp_path / "output.xlsx"

    result = CompareUseCase().execute(old, new, output, CompareOptions(json_array_key="id"))

    assert result.count(DifferenceType.MODIFIED) == 1
    difference = result.differences[0]
    assert difference.cell == '$.items[id="P002"].price'
    assert difference.old_value == 200
    assert difference.new_value == 250
    workbook = load_workbook(output)
    report = workbook["比較結果"]
    assert report["C11"].value == '$.items[id="P002"].price'
    assert report["D11"].value == 200
    assert report["E11"].value == 250
    assert report["H6"].value == "配列キー"
    assert report["I6"].value == "id"
    workbook.close()


def test_xml_files_are_compared_and_written_to_excel_report(tmp_path: Path) -> None:
    old = make_xml(tmp_path / "old.xml", '<root enabled="true"><name>old</name></root>')
    new = make_xml(tmp_path / "new.xml", '<root enabled="false"><name>new</name><item /></root>')
    output = tmp_path / "output.xlsx"

    result = CompareUseCase().execute(old, new, output, CompareOptions())

    assert result.count(DifferenceType.MODIFIED) == 2
    assert result.count(DifferenceType.ADDED) == 1
    workbook = load_workbook(output)
    assert workbook.sheetnames == ["比較結果", "XML"]
    report = workbook["比較結果"]
    rows = {report.cell(row, 3).value: row for row in range(11, 14)}
    assert report.cell(rows["/root/@enabled"], 1).value == "変更"
    assert report.cell(rows["/root/@enabled"], 2).value == "XML"
    assert report.cell(rows["/root/@enabled"], 4).value == "true"
    assert report.cell(rows["/root/@enabled"], 5).value == "false"
    assert report.cell(rows["/root/@enabled"], 6).value is None
    assert report["H1"].value == "XML読み込み設定"
    assert report["I4"].value == "はい"
    assert report["I5"].value == "はい"
    assert workbook["XML"]["A1"].value == "新XML"
    workbook.close()


def test_xml_options_are_used_and_recorded(tmp_path: Path) -> None:
    old = make_xml(tmp_path / "old.xml", '<root a="1" b="2"><name>A</name></root>')
    new = make_xml(tmp_path / "new.xml", '<root b="2" a="1"> <name>A</name></root>')
    output = tmp_path / "output.xlsx"

    result = CompareUseCase().execute(
        old,
        new,
        output,
        CompareOptions(ignore_xml_attribute_order=False, ignore_xml_blank_text=False),
    )

    assert result.count(DifferenceType.MODIFIED) == 2
    workbook = load_workbook(output)
    report = workbook["比較結果"]
    rows = {report.cell(row, 3).value: row for row in range(11, 13)}
    assert report.cell(rows["/root/@*"], 4).value == "a, b"
    assert report.cell(rows["/root/@*"], 5).value == "b, a"
    assert report.cell(rows["/root"], 5).value == " "
    assert report["I4"].value == "いいえ"
    assert report["I5"].value == "いいえ"
    workbook.close()


def test_xml_child_insert_is_reported_without_shifted_following_changes(tmp_path: Path) -> None:
    old = make_xml(tmp_path / "old.xml", "<root><item>A</item><item>B</item></root>")
    new = make_xml(tmp_path / "new.xml", "<root><item>A</item><item>X</item><item>B</item></root>")
    output = tmp_path / "output.xlsx"

    result = CompareUseCase().execute(old, new, output, CompareOptions())

    assert result.total == 1
    assert result.count(DifferenceType.ADDED) == 1
    workbook = load_workbook(output)
    report = workbook["比較結果"]
    assert report["A11"].value == "追加"
    assert report["B11"].value == "XML"
    assert report["C11"].value == "/root/item[2]"
    assert report["E11"].value == "<item>X</item>"
    assert report["A12"].value is None
    workbook.close()


def test_xml_children_are_matched_by_id_attribute_through_usecase(tmp_path: Path) -> None:
    old = make_xml(tmp_path / "old.xml", '<root><item id="P001" quantity="2" /><item id="P002" quantity="8" /></root>')
    new = make_xml(tmp_path / "new.xml", '<root><item id="P002" quantity="10" /><item id="P001" quantity="2" /></root>')
    output = tmp_path / "output.xlsx"

    result = CompareUseCase().execute(old, new, output, CompareOptions())

    assert result.total == 1
    assert result.count(DifferenceType.MODIFIED) == 1
    workbook = load_workbook(output)
    report = workbook["比較結果"]
    assert report["A11"].value == "変更"
    assert report["B11"].value == "XML"
    assert report["C11"].value == "/root/item[2]/@quantity"
    assert report["D11"].value == "8"
    assert report["E11"].value == "10"
    assert report["A12"].value is None
    workbook.close()


def test_xml_children_are_matched_by_configured_attribute_through_usecase(tmp_path: Path) -> None:
    old = make_xml(
        tmp_path / "old.xml",
        '<root><item code="P001" quantity="2" /><item code="P002" quantity="8" /></root>',
    )
    new = make_xml(
        tmp_path / "new.xml",
        '<root><item code="P002" quantity="10" /><item code="P001" quantity="2" /></root>',
    )
    output = tmp_path / "output.xlsx"

    result = CompareUseCase().execute(old, new, output, CompareOptions(xml_element_key_attribute="code"))

    assert result.total == 1
    assert result.count(DifferenceType.MODIFIED) == 1
    workbook = load_workbook(output)
    report = workbook["比較結果"]
    assert report["A11"].value == "変更"
    assert report["B11"].value == "XML"
    assert report["C11"].value == "/root/item[2]/@quantity"
    assert report["D11"].value == "8"
    assert report["E11"].value == "10"
    assert report["H6"].value == "要素キー属性"
    assert report["I6"].value == "code"
    assert report["A12"].value is None
    workbook.close()


def test_invalid_xml_is_reported_through_usecase(tmp_path: Path) -> None:
    old = make_xml(tmp_path / "old.xml", "<root>\n  <name>broken</root>")
    new = make_xml(tmp_path / "new.xml", "<root><name>new</name></root>")

    with pytest.raises(WorkbookReadError) as error:
        CompareUseCase().execute(old, new, tmp_path / "output.xlsx", CompareOptions())

    message = str(error.value)
    assert "XMLファイルの形式を読み取れません" in message
    assert "2行" in message
    assert "開始タグと終了タグ" in message


def test_xml_utf8_read_error_is_reported_through_usecase(tmp_path: Path) -> None:
    old = tmp_path / "old.xml"
    old.write_bytes(b"\xff\xfe\x00\x80")
    new = make_xml(tmp_path / "new.xml", "<root><name>new</name></root>")

    with pytest.raises(WorkbookReadError, match="UTF-8 / UTF-8 BOM付きで保存してください"):
        CompareUseCase().execute(old, new, tmp_path / "output.xlsx", CompareOptions())


def test_xml_summary_mode_omits_detail_table_but_keeps_counts(tmp_path: Path) -> None:
    old = make_xml(tmp_path / "old.xml", '<root enabled="true"><name>old</name></root>')
    new = make_xml(tmp_path / "new.xml", '<root enabled="false"><name>new</name><item /></root>')
    output = tmp_path / "output.xlsx"

    result = CompareUseCase().execute(old, new, output, CompareOptions(), detailed=False)

    assert result.count(DifferenceType.MODIFIED) == 2
    assert result.count(DifferenceType.ADDED) == 1
    workbook = load_workbook(output)
    report = workbook["比較結果"]
    assert report["A2"].value == "変更件数"
    assert report["B2"].value == 2
    assert report["A3"].value == "追加件数"
    assert report["B3"].value == 1
    assert report["A10"].value is None
    assert report["H1"].value == "XML読み込み設定"
    assert workbook["XML"]["A1"].value == "新XML"
    workbook.close()


def test_cancelled_xml_compare_does_not_create_output(tmp_path: Path) -> None:
    old = make_xml(tmp_path / "old.xml", "<root><name>old</name></root>")
    new = make_xml(tmp_path / "new.xml", "<root><name>new</name></root>")
    output = tmp_path / "output.xlsx"

    with pytest.raises(OperationCancelledError):
        CompareUseCase().execute(old, new, output, CompareOptions(), cancel_requested=lambda: True)

    assert not output.exists()


def test_missing_input_file_is_rejected(tmp_path: Path) -> None:
    new = make_workbook(tmp_path / "new.xlsx", {"Data": {}})
    with pytest.raises(InvalidInputError, match="見つかりません"):
        CompareUseCase().execute(tmp_path / "missing.xlsx", new, tmp_path / "output.xlsx", CompareOptions())


def test_input_path_cannot_be_used_as_output(tmp_path: Path) -> None:
    old = make_workbook(tmp_path / "old.xlsx", {"Data": {"A1": "old"}})
    new = make_workbook(tmp_path / "new.xlsx", {"Data": {"A1": "new"}})
    with pytest.raises(InvalidInputError, match="異なるパス"):
        CompareUseCase().execute(old, new, new, CompareOptions())


def test_corrupted_xlsx_is_reported_as_read_error(tmp_path: Path) -> None:
    corrupt = tmp_path / "corrupt.xlsx"
    corrupt.write_bytes(b"this is not an xlsx archive")
    new = make_workbook(tmp_path / "new.xlsx", {"Data": {}})
    with pytest.raises(WorkbookReadError, match="破損"):
        CompareUseCase().execute(corrupt, new, tmp_path / "output.xlsx", CompareOptions())


def test_unusable_output_parent_is_reported_as_write_error(tmp_path: Path) -> None:
    old = make_workbook(tmp_path / "old.xlsx", {"Data": {}})
    new = make_workbook(tmp_path / "new.xlsx", {"Data": {"A1": "new"}})
    parent_file = tmp_path / "not_a_directory"
    parent_file.write_text("file", encoding="utf-8")
    with pytest.raises(OutputWriteError, match="保存できません"):
        CompareUseCase().execute(old, new, parent_file / "output.xlsx", CompareOptions())


def test_existing_output_is_preserved_when_report_creation_fails(tmp_path: Path) -> None:
    class FailingWriter(ExcelReportWriter):
        def _write_report(
            self,
            sheet: object,
            result: CompareResult,
            detailed: bool,
            cancel_requested: object = None,
        ) -> None:
            raise ValueError("simulated report failure")

    source = make_workbook(tmp_path / "new.xlsx", {"Data": {"A1": "new"}})
    output = tmp_path / "existing.xlsx"
    original_content = b"existing output must survive"
    output.write_bytes(original_content)

    with pytest.raises(OutputWriteError):
        FailingWriter().write(source, output, CompareResult())

    assert output.read_bytes() == original_content
    assert list(tmp_path.glob(".existing_*.tmp.xlsx")) == []


def test_clearing_file_history_preserves_last_save_directory(tmp_path: Path) -> None:
    settings = AppSettingsStore(tmp_path / "settings.json")
    output = tmp_path / "output"
    output.mkdir()
    source = tmp_path / "source.xlsx"
    source.write_bytes(b"test")
    settings.save_last_save_dir(output)
    settings.save_file_history([str(source)])

    settings.clear_file_history()

    assert settings.load_file_history() == []
    assert settings.load_last_save_dir() == output.resolve()


def test_compare_options_and_view_mode_are_persisted(tmp_path: Path) -> None:
    settings = AppSettingsStore(tmp_path / "settings.json")
    options = CompareOptions(
        compare_values=False,
        compare_formulas=False,
        empty_string_equals_empty=False,
        ignore_surrounding_whitespace=True,
        ignore_case=True,
        algorithm=CompareAlgorithm.KEY_COLUMNS,
        key_columns=("A", "C"),
        csv_encoding="cp932",
        csv_delimiter="\t",
        ignore_csv_blank_lines=False,
        ignore_json_object_key_order=False,
        ignore_json_array_order=True,
        json_array_key="id",
        ignore_xml_attribute_order=False,
        ignore_xml_blank_text=False,
        xml_element_key_attribute="code",
    )

    settings.save_compare_options(options)
    settings.save_view_mode("summary")

    loaded = settings.load_compare_options()
    assert loaded == options
    assert settings.load_view_mode() == "summary"


def test_successful_report_atomically_replaces_existing_output(tmp_path: Path) -> None:
    old = make_workbook(tmp_path / "old.xlsx", {"Data": {"A1": "old"}})
    new = make_workbook(tmp_path / "new.xlsx", {"Data": {"A1": "new"}})
    output = tmp_path / "existing.xlsx"
    output.write_bytes(b"previous output")

    CompareUseCase().execute(old, new, output, CompareOptions())

    workbook = load_workbook(output)
    assert workbook.sheetnames[0] == "比較結果"
    assert workbook["Data"]["A1"].value == "new"
    workbook.close()
    assert list(tmp_path.glob(".existing_*.tmp.xlsx")) == []


def test_cancelled_compare_does_not_create_output(tmp_path: Path) -> None:
    old = make_workbook(tmp_path / "old.xlsx", {"Data": {"A1": "old"}})
    new = make_workbook(tmp_path / "new.xlsx", {"Data": {"A1": "new"}})
    output = tmp_path / "output.xlsx"

    with pytest.raises(OperationCancelledError):
        CompareUseCase().execute(old, new, output, CompareOptions(), cancel_requested=lambda: True)

    assert not output.exists()


def test_cancelled_report_preserves_existing_output(tmp_path: Path) -> None:
    source = make_workbook(tmp_path / "new.xlsx", {"Data": {"A1": "new"}})
    output = tmp_path / "existing.xlsx"
    original_content = b"existing output must survive cancellation"
    output.write_bytes(original_content)
    result = CompareResult(
        [
            Difference(DifferenceType.MODIFIED, "Data", "A1", "old", "new"),
            Difference(DifferenceType.ADDED, "Data", "B1", None, "added"),
        ]
    )

    with pytest.raises(OperationCancelledError):
        ExcelReportWriter().write(source, output, result, cancel_requested=lambda: True)

    assert output.read_bytes() == original_content
    assert list(tmp_path.glob(".existing_*.tmp.xlsx")) == []


def test_existing_report_sheet_names_are_not_overwritten(tmp_path: Path) -> None:
    sheets = {"Data": {"A1": "old"}, "比較結果": {"A1": "keep"}, "比較結果_1": {"A1": "keep too"}}
    new_sheets = {**sheets, "Data": {"A1": "new"}}
    _, output = compare(tmp_path, sheets, new_sheets)
    workbook = load_workbook(output)
    assert workbook.sheetnames[:3] == ["比較結果_2", "Data", "比較結果"]
    assert workbook["比較結果"]["A1"].value == "keep"
    assert workbook["比較結果_1"]["A1"].value == "keep too"
    workbook.close()


def test_link_escapes_apostrophe_in_sheet_name(tmp_path: Path) -> None:
    sheet_name = "社長's Data"
    _, output = compare(tmp_path, {sheet_name: {"A1": "old"}}, {sheet_name: {"A1": "new"}})
    workbook = load_workbook(output)
    assert workbook["比較結果"]["F11"].hyperlink.target == "#'社長''s Data'!A1"
    workbook.close()


def test_summary_mode_omits_detail_table_but_keeps_counts(tmp_path: Path) -> None:
    result, output = compare(tmp_path, {"Data": {"A1": "old"}}, {"Data": {"A1": "new", "B1": "added"}}, detailed=False)
    workbook = load_workbook(output)
    report = workbook["比較結果"]
    assert report["B2"].value == result.count(DifferenceType.MODIFIED) == 1
    assert report["B3"].value == result.count(DifferenceType.ADDED) == 1
    assert report["A10"].value is None
    workbook.close()


def test_row_lcs_report_summarizes_inserted_row(tmp_path: Path) -> None:
    result, output = compare(
        tmp_path,
        {"Data": {"A1": "A", "A2": "B", "A3": "C"}},
        {"Data": {"A1": "A", "A2": "X", "A3": "B", "A4": "C"}},
        options=CompareOptions(algorithm=CompareAlgorithm.ROW_LCS),
    )

    assert result.count(DifferenceType.ROW_ADDED) == 1
    assert result.count(DifferenceType.MODIFIED) == 0
    workbook = load_workbook(output)
    report = workbook["比較結果"]
    assert report["B5"].value == 1
    assert report["A11"].value == "行追加"
    assert report["C11"].value == "2:2"
    assert report["F11"].value is None
    workbook.close()


def test_key_column_report_matches_moved_row_by_key(tmp_path: Path) -> None:
    result, output = compare(
        tmp_path,
        {"Data": {"A1": "ID", "B1": "Value", "A2": "001", "B2": 10}},
        {"Data": {"A1": "ID", "B1": "Value", "A5": "001", "B5": 20}},
        options=CompareOptions(algorithm=CompareAlgorithm.KEY_COLUMNS, key_columns=("A",)),
    )

    assert result.count(DifferenceType.MODIFIED) == 1
    assert result.count(DifferenceType.ROW_ADDED) == 0
    assert result.count(DifferenceType.ROW_DELETED) == 0
    workbook = load_workbook(output)
    report = workbook["比較結果"]
    assert report["A11"].value == "変更"
    assert report["C11"].value == "B5"
    assert report["D11"].value == 10
    assert report["E11"].value == 20
    workbook.close()


def test_large_detailed_compare_reports_output_notice(tmp_path: Path) -> None:
    old_cells = {f"A{row}": f"old {row}" for row in range(1, 1001)}
    new_cells = {f"A{row}": f"new {row}" for row in range(1, 1001)}
    old = make_workbook(tmp_path / "old.xlsx", {"Data": old_cells})
    new = make_workbook(tmp_path / "new.xlsx", {"Data": new_cells})
    messages: list[str] = []

    result = CompareUseCase().execute(
        old,
        new,
        tmp_path / "output.xlsx",
        CompareOptions(),
        detailed=True,
        progress_callback=messages.append,
    )

    assert result.total == 1000
    assert "差分を 1,000 件検出しました。" in messages
    assert "差分が多いため、詳細レポートの作成に時間がかかる場合があります。" in messages
    assert messages[-1] == "比較結果Excelの作成が完了しました。"


def test_large_summary_compare_does_not_report_detailed_notice(tmp_path: Path) -> None:
    old_cells = {f"A{row}": f"old {row}" for row in range(1, 1001)}
    new_cells = {f"A{row}": f"new {row}" for row in range(1, 1001)}
    old = make_workbook(tmp_path / "old.xlsx", {"Data": old_cells})
    new = make_workbook(tmp_path / "new.xlsx", {"Data": new_cells})
    messages: list[str] = []

    CompareUseCase().execute(
        old,
        new,
        tmp_path / "output.xlsx",
        CompareOptions(),
        detailed=False,
        progress_callback=messages.append,
    )

    assert "差分を 1,000 件検出しました。" in messages
    assert "差分が多いため、詳細レポートの作成に時間がかかる場合があります。" not in messages


def test_formula_text_change_is_reported_without_recalculation(tmp_path: Path) -> None:
    result, output = compare(
        tmp_path,
        {"Data": {"A1": 1, "A2": 2, "B1": "=SUM(A1:A2)"}},
        {"Data": {"A1": 1, "A2": 2, "B1": "=SUM(A1:A1)"}},
    )
    difference = result.differences[0]
    assert difference.cell == "B1"
    assert difference.formula_changed is True
    workbook = load_workbook(output)
    assert workbook["比較結果"]["D11"].value == "'=SUM(A1:A2)"
    assert workbook["比較結果"]["E11"].value == "'=SUM(A1:A1)"
    workbook.close()


def test_formatting_only_cells_are_ignored(tmp_path: Path) -> None:
    old = make_workbook(tmp_path / "old.xlsx", {"Data": {"A1": "same"}})
    new = make_workbook(tmp_path / "new.xlsx", {"Data": {"A1": "same"}})
    workbook = load_workbook(new)
    workbook["Data"]["Z100"].fill = PatternFill("solid", fgColor="FF0000")
    workbook.save(new)
    workbook.close()
    result = CompareUseCase().execute(old, new, tmp_path / "output.xlsx", CompareOptions())
    assert result.total == 0


def test_sheet_addition_and_deletion_are_summarized(tmp_path: Path) -> None:
    result, output = compare(tmp_path, {"Common": {}, "Old only": {}}, {"Common": {}, "New only": {}})
    assert result.count(DifferenceType.SHEET_ADDED) == 1
    assert result.count(DifferenceType.SHEET_DELETED) == 1
    workbook = load_workbook(output)
    report = workbook["比較結果"]
    assert report["B7"].value == 1
    assert report["B8"].value == 1
    workbook.close()
