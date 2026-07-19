from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from compare_tool.errors import OperationCancelledError, WorkbookReadError
from compare_tool.models import CompareOptions, DifferenceType
from compare_tool.xml_compare import XML_SHEET_NAME, XmlComparer, XmlReader, XmlReportWriter


def write_xml(path: Path, text: str, *, encoding: str = "utf-8") -> Path:
    path.write_text(text, encoding=encoding)
    return path


def test_xml_reader_reads_utf8_bom_file(tmp_path: Path) -> None:
    path = tmp_path / "input.xml"
    path.write_text("\ufeff<root><name>compare_tool</name></root>", encoding="utf-8")

    document = XmlReader().read(path)
    name = document.root.find("name")

    assert document.root.tag == "root"
    assert name is not None
    assert name.text == "compare_tool"


def test_xml_reader_reports_invalid_xml(tmp_path: Path) -> None:
    path = write_xml(tmp_path / "broken.xml", "<root>\n  <name>broken</root>")

    with pytest.raises(WorkbookReadError) as error:
        XmlReader().read(path)
    message = str(error.value)
    assert "XMLファイルの形式を読み取れません" in message
    assert "2行" in message
    assert "開始タグと終了タグ" in message


def test_xml_comparer_reports_text_attribute_and_child_differences(tmp_path: Path) -> None:
    old = XmlReader().read(
        write_xml(
            tmp_path / "old.xml",
            '<root enabled="true" removed="yes"><name>old</name><obsolete /></root>',
        )
    )
    new = XmlReader().read(
        write_xml(
            tmp_path / "new.xml",
            '<root enabled="false" added="yes"><name>new</name><item /></root>',
        )
    )

    result = XmlComparer().compare(old, new, CompareOptions())

    differences = {
        (difference.kind, difference.cell, difference.old_value, difference.new_value)
        for difference in result.differences
    }
    assert differences == {
        (DifferenceType.ADDED, "/root/@added", None, "yes"),
        (DifferenceType.ADDED, "/root/item[1]", None, "<item />"),
        (DifferenceType.DELETED, "/root/@removed", "yes", None),
        (DifferenceType.DELETED, "/root/obsolete[1]", "<obsolete />", None),
        (DifferenceType.MODIFIED, "/root/@enabled", "true", "false"),
        (DifferenceType.MODIFIED, "/root/name[1]", "old", "new"),
    }
    assert {difference.sheet for difference in result.differences} == {XML_SHEET_NAME}


def test_xml_comparer_reports_added_and_deleted_children(tmp_path: Path) -> None:
    old = XmlReader().read(write_xml(tmp_path / "old.xml", "<root><item>A</item><item>B</item></root>"))
    new = XmlReader().read(write_xml(tmp_path / "new.xml", "<root><item>A</item><item>B</item><item>C</item></root>"))

    result = XmlComparer().compare(old, new, CompareOptions())

    assert result.total == 1
    difference = result.differences[0]
    assert difference.kind is DifferenceType.ADDED
    assert difference.cell == "/root/item[3]"
    assert difference.new_value == "<item>C</item>"


def test_xml_comparer_uses_lcs_to_avoid_shifted_child_differences(tmp_path: Path) -> None:
    old = XmlReader().read(write_xml(tmp_path / "old.xml", "<root><item>A</item><item>B</item></root>"))
    new = XmlReader().read(write_xml(tmp_path / "new.xml", "<root><item>A</item><item>X</item><item>B</item></root>"))

    result = XmlComparer().compare(old, new, CompareOptions())

    assert result.total == 1
    difference = result.differences[0]
    assert difference.kind is DifferenceType.ADDED
    assert difference.cell == "/root/item[2]"
    assert difference.new_value == "<item>X</item>"


def test_xml_comparer_ignores_blank_text_by_default(tmp_path: Path) -> None:
    old = XmlReader().read(write_xml(tmp_path / "old.xml", "<root><item>A</item></root>"))
    new = XmlReader().read(
        write_xml(
            tmp_path / "new.xml",
            """
            <root>
              <item>A</item>
            </root>
            """,
        )
    )

    result = XmlComparer().compare(old, new, CompareOptions())

    assert result.total == 0


def test_xml_comparer_can_compare_blank_text(tmp_path: Path) -> None:
    old = XmlReader().read(write_xml(tmp_path / "old.xml", "<root><item>A</item></root>"))
    new = XmlReader().read(write_xml(tmp_path / "new.xml", "<root> <item>A</item></root>"))

    result = XmlComparer().compare(old, new, CompareOptions(ignore_xml_blank_text=False))

    assert result.total == 1
    assert result.differences[0].cell == "/root"
    assert result.differences[0].new_value == " "


def test_xml_comparer_can_compare_attribute_order(tmp_path: Path) -> None:
    old = XmlReader().read(write_xml(tmp_path / "old.xml", '<root a="1" b="2" />'))
    new = XmlReader().read(write_xml(tmp_path / "new.xml", '<root b="2" a="1" />'))

    ignored = XmlComparer().compare(old, new, CompareOptions())
    compared = XmlComparer().compare(old, new, CompareOptions(ignore_xml_attribute_order=False))

    assert ignored.total == 0
    assert compared.total == 1
    assert compared.differences[0].cell == "/root/@*"
    assert compared.differences[0].old_value == ["a", "b"]
    assert compared.differences[0].new_value == ["b", "a"]


def test_xml_comparer_uses_string_normalization_options(tmp_path: Path) -> None:
    old = XmlReader().read(write_xml(tmp_path / "old.xml", "<root><name> Sample </name></root>"))
    new = XmlReader().read(write_xml(tmp_path / "new.xml", "<root><name>sample</name></root>"))

    result = XmlComparer().compare(
        old,
        new,
        CompareOptions(ignore_surrounding_whitespace=True, ignore_case=True),
    )

    assert result.total == 0


def test_xml_comparer_can_be_cancelled(tmp_path: Path) -> None:
    old = XmlReader().read(write_xml(tmp_path / "old.xml", "<root><name>old</name></root>"))
    new = XmlReader().read(write_xml(tmp_path / "new.xml", "<root><name>new</name></root>"))

    with pytest.raises(OperationCancelledError):
        XmlComparer().compare(old, new, CompareOptions(), cancel_requested=lambda: True)


def test_xml_report_writer_outputs_xlsx_report(tmp_path: Path) -> None:
    old_path = write_xml(tmp_path / "old.xml", '<root enabled="true"><name>old</name></root>')
    new_path = write_xml(tmp_path / "new.xml", '<root enabled="false"><name>new</name><item /></root>')
    old = XmlReader().read(old_path)
    new = XmlReader().read(new_path)
    options = CompareOptions(ignore_xml_attribute_order=False, ignore_xml_blank_text=False)
    result = XmlComparer().compare(old, new, options)
    output = tmp_path / "output.xlsx"

    XmlReportWriter().write(new_path, output, result, options=options)

    workbook = load_workbook(output)
    assert workbook.sheetnames == ["比較結果", "XML"]
    report = workbook["比較結果"]
    rows = {report.cell(row, 3).value: row for row in range(11, 14)}
    assert report.cell(rows["/root/@enabled"], 1).value == "変更"
    assert report.cell(rows["/root/@enabled"], 2).value == "XML"
    assert report.cell(rows["/root/@enabled"], 4).value == "true"
    assert report.cell(rows["/root/@enabled"], 5).value == "false"
    assert report.cell(rows["/root/@enabled"], 6).value is None
    assert report["H1"].value == "XML読み込み設定"
    assert report["I2"].value == "UTF-8 / UTF-8 BOM"
    assert report["I3"].value == "XPath風パス"
    assert report["I4"].value == "いいえ"
    assert report["I5"].value == "いいえ"
    xml_sheet = workbook["XML"]
    assert xml_sheet["A1"].value == "新XML"
    assert xml_sheet["A2"].value == '<root enabled="false">'
    assert xml_sheet["A3"].value == "  <name>new</name>"
    workbook.close()
