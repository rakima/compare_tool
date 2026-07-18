"""Opt-in performance regression tests for large Excel workbooks."""

from __future__ import annotations

import time
import tracemalloc
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

from compare_tool.models import CompareOptions, DifferenceType
from compare_tool.usecase import CompareUseCase

pytestmark = pytest.mark.performance


def create_dense_workbook(path: Path, rows: int = 5_000, columns: int = 10, changed: bool = False) -> None:
    workbook = Workbook(write_only=True)
    sheet = workbook.create_sheet("Data")
    for row_number in range(1, rows + 1):
        values: list[object] = [row_number * 100 + column for column in range(1, columns + 1)]
        if changed and row_number == rows:
            values[-1] = "changed"
        sheet.append(values)
    workbook.save(path)
    workbook.close()


def create_shifted_workbook(path: Path, rows: int = 1_000, columns: int = 10, offset: int = 0) -> None:
    workbook = Workbook(write_only=True)
    sheet = workbook.create_sheet("Data")
    for row_number in range(1, rows + 1):
        sheet.append([row_number * 100 + column + offset for column in range(1, columns + 1)])
    workbook.save(path)
    workbook.close()


def test_compare_50k_cells_with_bounded_time_and_memory(tmp_path: Path) -> None:
    old = tmp_path / "old.xlsx"
    new = tmp_path / "new.xlsx"
    output = tmp_path / "output.xlsx"
    create_dense_workbook(old)
    create_dense_workbook(new, changed=True)

    tracemalloc.start()
    started = time.perf_counter()
    result = CompareUseCase().execute(old, new, output, CompareOptions())
    elapsed = time.perf_counter() - started
    _, peak_bytes = tracemalloc.get_traced_memory()
    tracemalloc.stop()

    assert result.count(DifferenceType.MODIFIED) == 1
    assert output.is_file()
    assert elapsed < 30, f"50,000-cell comparison took {elapsed:.2f}s"
    assert peak_bytes < 350 * 1024 * 1024, f"peak memory was {peak_bytes / 1024 / 1024:.1f} MiB"


def test_distant_formatting_does_not_expand_comparison_scan(tmp_path: Path) -> None:
    old = tmp_path / "old.xlsx"
    new = tmp_path / "new.xlsx"
    for path in (old, new):
        workbook = Workbook()
        sheet = workbook.active
        sheet["A1"] = "same"
        sheet["XFD1048576"].fill = PatternFill("solid", fgColor="FF0000")
        workbook.save(path)
        workbook.close()

    started = time.perf_counter()
    result = CompareUseCase().execute(old, new, tmp_path / "output.xlsx", CompareOptions())
    elapsed = time.perf_counter() - started

    assert result.total == 0
    assert elapsed < 10, f"sparse maximum-dimension workbook took {elapsed:.2f}s"


def test_detailed_report_with_10k_differences_has_bounded_runtime(tmp_path: Path) -> None:
    old = tmp_path / "old.xlsx"
    new = tmp_path / "new.xlsx"
    output = tmp_path / "output.xlsx"
    create_shifted_workbook(old)
    create_shifted_workbook(new, offset=1_000_000)

    started = time.perf_counter()
    result = CompareUseCase().execute(old, new, output, CompareOptions(), detailed=True)
    elapsed = time.perf_counter() - started

    assert result.count(DifferenceType.MODIFIED) == 10_000
    assert output.is_file()
    assert elapsed < 35, f"detailed report for 10,000 differences took {elapsed:.2f}s"


def test_summary_report_with_10k_differences_omits_detail_rows(tmp_path: Path) -> None:
    old = tmp_path / "old.xlsx"
    new = tmp_path / "new.xlsx"
    output = tmp_path / "output.xlsx"
    create_shifted_workbook(old)
    create_shifted_workbook(new, offset=1_000_000)

    started = time.perf_counter()
    result = CompareUseCase().execute(old, new, output, CompareOptions(), detailed=False)
    elapsed = time.perf_counter() - started

    workbook = load_workbook(output, read_only=True)
    try:
        report = workbook["比較結果"]
        assert report["B2"].value == 10_000
        assert report["A10"].value is None
    finally:
        workbook.close()
    assert result.count(DifferenceType.MODIFIED) == 10_000
    assert elapsed < 20, f"summary report for 10,000 differences took {elapsed:.2f}s"
