from __future__ import annotations

import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

from compare_tool.errors import InvalidInputError, OperationCancelledError, WorkbookReadError
from compare_tool.json_compare import JSON_SHEET_NAME, JsonComparer, JsonReader, JsonReportWriter
from compare_tool.models import CompareOptions, DifferenceType


def write_json(path: Path, data: object) -> Path:
    path.write_text(json.dumps(data, ensure_ascii=False), encoding="utf-8")
    return path


def test_json_reader_reads_utf8_bom_file(tmp_path: Path) -> None:
    path = tmp_path / "input.json"
    path.write_text('\ufeff{"name": "compare_tool"}', encoding="utf-8")

    document = JsonReader().read(path)

    assert document.data == {"name": "compare_tool"}


def test_json_reader_reports_invalid_json(tmp_path: Path) -> None:
    path = tmp_path / "broken.json"
    path.write_text('{\n  "name": "broken",\n}', encoding="utf-8")

    with pytest.raises(WorkbookReadError) as error:
        JsonReader().read(path)
    message = str(error.value)
    assert "3行 1列付近" in message
    assert "カンマ、引用符、コロン、括弧" in message


def test_json_comparer_reports_modified_added_and_deleted_paths(tmp_path: Path) -> None:
    old = JsonReader().read(
        write_json(tmp_path / "old.json", {"name": "old", "removed": True, "items": [{"price": 100}]})
    )
    new = JsonReader().read(write_json(tmp_path / "new.json", {"name": "new", "added": 1, "items": [{"price": 120}]}))

    result = JsonComparer().compare(old, new, CompareOptions())

    differences = sorted((difference.cell, difference.kind) for difference in result.differences)
    assert differences == [
        ("$.added", DifferenceType.ADDED),
        ("$.items[0].price", DifferenceType.MODIFIED),
        ("$.name", DifferenceType.MODIFIED),
        ("$.removed", DifferenceType.DELETED),
    ]
    assert {difference.sheet for difference in result.differences} == {JSON_SHEET_NAME}


def test_json_comparer_reports_array_items_by_index(tmp_path: Path) -> None:
    old = JsonReader().read(write_json(tmp_path / "old.json", ["A", "B"]))
    new = JsonReader().read(write_json(tmp_path / "new.json", ["A", "X", "C"]))

    result = JsonComparer().compare(old, new, CompareOptions())

    differences = [
        (difference.kind, difference.cell, difference.old_value, difference.new_value)
        for difference in result.differences
    ]
    assert differences == [
        (DifferenceType.MODIFIED, "$[1]", "B", "X"),
        (DifferenceType.ADDED, "$[2]", None, "C"),
    ]


def test_json_comparer_uses_string_normalization_options(tmp_path: Path) -> None:
    old = JsonReader().read(write_json(tmp_path / "old.json", {"name": " Sample "}))
    new = JsonReader().read(write_json(tmp_path / "new.json", {"name": "sample"}))

    result = JsonComparer().compare(
        old,
        new,
        CompareOptions(ignore_surrounding_whitespace=True, ignore_case=True),
    )

    assert result.total == 0


def test_json_comparer_can_compare_object_key_order(tmp_path: Path) -> None:
    old = JsonReader().read(write_json(tmp_path / "old.json", {"a": 1, "b": 2}))
    new = JsonReader().read(write_json(tmp_path / "new.json", {"b": 2, "a": 1}))

    ignored = JsonComparer().compare(old, new, CompareOptions())
    compared = JsonComparer().compare(old, new, CompareOptions(ignore_json_object_key_order=False))

    assert ignored.total == 0
    assert compared.total == 1
    difference = compared.differences[0]
    assert difference.kind is DifferenceType.MODIFIED
    assert difference.cell == "$"
    assert difference.old_value == ["a", "b"]
    assert difference.new_value == ["b", "a"]


def test_json_comparer_can_ignore_array_order(tmp_path: Path) -> None:
    old = JsonReader().read(write_json(tmp_path / "old.json", {"items": [{"id": 1}, {"id": 2}]}))
    new = JsonReader().read(write_json(tmp_path / "new.json", {"items": [{"id": 2}, {"id": 1}]}))

    indexed = JsonComparer().compare(old, new, CompareOptions())
    ignored = JsonComparer().compare(old, new, CompareOptions(ignore_json_array_order=True))

    assert indexed.total == 2
    assert ignored.total == 0


def test_json_comparer_reports_unmatched_array_items_when_ignoring_order(tmp_path: Path) -> None:
    old = JsonReader().read(write_json(tmp_path / "old.json", {"items": ["same", "same", "old"]}))
    new = JsonReader().read(write_json(tmp_path / "new.json", {"items": ["same", "new", "same"]}))

    result = JsonComparer().compare(old, new, CompareOptions(ignore_json_array_order=True))

    differences = [
        (difference.kind, difference.cell, difference.old_value, difference.new_value)
        for difference in result.differences
    ]
    assert differences == [
        (DifferenceType.DELETED, "$.items[2]", "old", None),
        (DifferenceType.ADDED, "$.items[1]", None, "new"),
    ]


def test_json_comparer_matches_array_objects_by_key(tmp_path: Path) -> None:
    old = JsonReader().read(
        write_json(
            tmp_path / "old.json",
            {"items": [{"id": "P001", "price": 100}, {"id": "P002", "price": 200}]},
        )
    )
    new = JsonReader().read(
        write_json(
            tmp_path / "new.json",
            {"items": [{"id": "P002", "price": 250}, {"id": "P001", "price": 100}]},
        )
    )

    result = JsonComparer().compare(old, new, CompareOptions(json_array_key="id"))

    assert [
        (difference.kind, difference.cell, difference.old_value, difference.new_value)
        for difference in result.differences
    ] == [(DifferenceType.MODIFIED, '$.items[id="P002"].price', 200, 250)]


def test_json_comparer_reports_keyed_array_added_and_deleted_items(tmp_path: Path) -> None:
    old = JsonReader().read(write_json(tmp_path / "old.json", {"items": [{"id": "P001"}, {"id": "P002"}]}))
    new = JsonReader().read(write_json(tmp_path / "new.json", {"items": [{"id": "P002"}, {"id": "P003"}]}))

    result = JsonComparer().compare(old, new, CompareOptions(json_array_key="id"))

    assert [(difference.kind, difference.cell) for difference in result.differences] == [
        (DifferenceType.DELETED, '$.items[id="P001"]'),
        (DifferenceType.ADDED, '$.items[id="P003"]'),
    ]


def test_json_comparer_rejects_duplicate_array_key_values(tmp_path: Path) -> None:
    old = JsonReader().read(write_json(tmp_path / "old.json", {"items": [{"id": "P001"}, {"id": "P001"}]}))
    new = JsonReader().read(write_json(tmp_path / "new.json", {"items": [{"id": "P001"}]}))

    with pytest.raises(InvalidInputError, match="JSON配列キー `id` の値が重複"):
        JsonComparer().compare(old, new, CompareOptions(json_array_key="id"))


def test_json_comparer_escapes_non_identifier_keys(tmp_path: Path) -> None:
    old = JsonReader().read(write_json(tmp_path / "old.json", {"display name": "old"}))
    new = JsonReader().read(write_json(tmp_path / "new.json", {"display name": "new"}))

    result = JsonComparer().compare(old, new, CompareOptions())

    assert result.differences[0].cell == "$['display name']"


def test_json_comparer_can_be_cancelled(tmp_path: Path) -> None:
    old = JsonReader().read(write_json(tmp_path / "old.json", {"name": "old"}))
    new = JsonReader().read(write_json(tmp_path / "new.json", {"name": "new"}))

    with pytest.raises(OperationCancelledError):
        JsonComparer().compare(old, new, CompareOptions(), cancel_requested=lambda: True)


def test_json_report_writer_outputs_xlsx_report(tmp_path: Path) -> None:
    old_path = write_json(tmp_path / "old.json", {"name": "old"})
    new_path = write_json(tmp_path / "new.json", {"name": "new", "items": [1, 2]})
    old = JsonReader().read(old_path)
    new = JsonReader().read(new_path)
    result = JsonComparer().compare(old, new, CompareOptions())
    output = tmp_path / "output.xlsx"

    JsonReportWriter().write(
        new_path,
        output,
        result,
        options=CompareOptions(ignore_json_object_key_order=False, ignore_json_array_order=True, json_array_key="id"),
    )

    workbook = load_workbook(output)
    assert workbook.sheetnames == ["比較結果", "JSON"]
    report = workbook["比較結果"]
    assert report["A11"].value == "追加"
    assert report["B11"].value == "JSON"
    assert report["C11"].value == "$.items"
    assert report["F11"].value is None
    assert report["H1"].value == "JSON読み込み設定"
    assert report["I2"].value == "UTF-8 / UTF-8 BOM"
    assert report["I3"].value == "JSON Path"
    assert report["I4"].value == "いいえ"
    assert report["I5"].value == "はい"
    assert report["I6"].value == "id"
    json_sheet = workbook["JSON"]
    assert json_sheet["A1"].value == "新JSON"
    assert json_sheet["A2"].value == "{"
    assert '"name": "new"' in json_sheet["A3"].value
    workbook.close()
