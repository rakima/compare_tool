from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from compare_tool.errors import InvalidInputError, PasswordProtectedWorkbookError
from compare_tool.excel import CellCoordinateCompareAlgorithm, CellData, ExcelComparer, ExcelDocument
from compare_tool.models import CompareAlgorithm, CompareOptions, Difference, DifferenceType
from compare_tool.usecase import CompareUseCase


def test_comparer_detects_cells_and_sheets() -> None:
    old = ExcelDocument({"共通": {"A1": CellData("Old"), "A2": CellData("gone")}, "削除": {}})
    new = ExcelDocument({"共通": {"A1": CellData("New"), "B1": CellData("added")}, "追加": {}})
    result = ExcelComparer().compare(old, new, CompareOptions())
    assert [(d.kind, d.sheet, d.cell) for d in result.differences] == [
        (DifferenceType.SHEET_ADDED, "追加", None),
        (DifferenceType.SHEET_DELETED, "削除", None),
        (DifferenceType.MODIFIED, "共通", "A1"),
        (DifferenceType.DELETED, "共通", "A2"),
        (DifferenceType.ADDED, "共通", "B1"),
    ]


def test_normalization_options() -> None:
    old = ExcelDocument({"S": {"A1": CellData(" Value "), "A2": CellData("")}})
    new = ExcelDocument({"S": {"A1": CellData("value")}})
    options = CompareOptions(ignore_surrounding_whitespace=True, ignore_case=True, empty_string_equals_empty=True)
    assert ExcelComparer().compare(old, new, options).total == 0


def test_formula_comparison_can_be_disabled() -> None:
    old = ExcelDocument({"S": {"A1": CellData(10, "=SUM(B1:B2)")}})
    new = ExcelDocument({"S": {"A1": CellData(10, "=SUM(B1:B3)")}})
    assert ExcelComparer().compare(old, new, CompareOptions(compare_formulas=False)).total == 0
    difference = ExcelComparer().compare(old, new, CompareOptions()).differences[0]
    assert difference.formula_changed
    assert difference.old_value == "=SUM(B1:B2)"


def test_excel_comparer_uses_selected_algorithm() -> None:
    class StubAlgorithm(CellCoordinateCompareAlgorithm):
        def compare_sheet(
            self,
            sheet: str,
            old_cells: dict[str, CellData],
            new_cells: dict[str, CellData],
            options: CompareOptions,
            cancel_requested: object = None,
        ) -> list[Difference]:
            return [Difference(DifferenceType.MODIFIED, sheet, "A1", "stub-old", "stub-new")]

    comparer = ExcelComparer({CompareAlgorithm.CELL_COORDINATE: StubAlgorithm()})
    result = comparer.compare(ExcelDocument({"S": {}}), ExcelDocument({"S": {}}), CompareOptions())

    assert [(difference.kind, difference.sheet, difference.cell) for difference in result.differences] == [
        (DifferenceType.MODIFIED, "S", "A1")
    ]


def test_row_lcs_detects_inserted_row_without_shifting_following_rows() -> None:
    old = ExcelDocument({"S": {"A1": CellData("A"), "A2": CellData("B"), "A3": CellData("C")}})
    new = ExcelDocument({"S": {"A1": CellData("A"), "A2": CellData("X"), "A3": CellData("B"), "A4": CellData("C")}})
    result = ExcelComparer().compare(old, new, CompareOptions(algorithm=CompareAlgorithm.ROW_LCS))

    assert [(difference.kind, difference.cell) for difference in result.differences] == [
        (DifferenceType.ROW_ADDED, "2:2")
    ]


def test_row_lcs_detects_deleted_row_without_shifting_following_rows() -> None:
    old = ExcelDocument({"S": {"A1": CellData("A"), "A2": CellData("B"), "A3": CellData("C")}})
    new = ExcelDocument({"S": {"A1": CellData("A"), "A2": CellData("C")}})
    result = ExcelComparer().compare(old, new, CompareOptions(algorithm=CompareAlgorithm.ROW_LCS))

    assert [(difference.kind, difference.cell) for difference in result.differences] == [
        (DifferenceType.ROW_DELETED, "2:2")
    ]


def test_key_column_compare_matches_rows_by_key_and_reports_cell_change() -> None:
    old = ExcelDocument(
        {"S": {"A1": CellData("ID"), "B1": CellData("Value"), "A2": CellData("001"), "B2": CellData(10)}}
    )
    new = ExcelDocument(
        {
            "S": {
                "A1": CellData("ID"),
                "B1": CellData("Value"),
                "A5": CellData("001"),
                "B5": CellData(20),
            }
        }
    )
    result = ExcelComparer().compare(
        old,
        new,
        CompareOptions(algorithm=CompareAlgorithm.KEY_COLUMNS, key_columns=("A",)),
    )

    actual = [
        (difference.kind, difference.cell, difference.old_value, difference.new_value)
        for difference in result.differences
    ]
    assert actual == [(DifferenceType.MODIFIED, "B5", 10, 20)]


def test_key_column_compare_reports_added_and_deleted_keys() -> None:
    old = ExcelDocument({"S": {"A2": CellData("001"), "B2": CellData("old")}})
    new = ExcelDocument({"S": {"A3": CellData("002"), "B3": CellData("new")}})
    result = ExcelComparer().compare(
        old,
        new,
        CompareOptions(algorithm=CompareAlgorithm.KEY_COLUMNS, key_columns=("A",)),
    )

    assert [(difference.kind, difference.cell) for difference in result.differences] == [
        (DifferenceType.ROW_DELETED, "2:2"),
        (DifferenceType.ROW_ADDED, "3:3"),
    ]


def test_key_column_compare_rejects_duplicate_keys() -> None:
    old = ExcelDocument({"S": {"A2": CellData("001"), "A3": CellData("001")}})
    new = ExcelDocument({"S": {"A2": CellData("001")}})

    with pytest.raises(InvalidInputError, match="重複"):
        ExcelComparer().compare(
            old,
            new,
            CompareOptions(algorithm=CompareAlgorithm.KEY_COLUMNS, key_columns=("A",)),
        )


def test_key_column_compare_rejects_cell_reference_as_key_column() -> None:
    old = ExcelDocument({"S": {"A2": CellData("001")}})
    new = ExcelDocument({"S": {"A2": CellData("001")}})

    with pytest.raises(InvalidInputError, match="列名だけ"):
        ExcelComparer().compare(
            old,
            new,
            CompareOptions(algorithm=CompareAlgorithm.KEY_COLUMNS, key_columns=("A1",)),
        )


def test_key_column_compare_rejects_repeated_key_columns() -> None:
    old = ExcelDocument({"S": {"A2": CellData("001")}})
    new = ExcelDocument({"S": {"A2": CellData("001")}})

    with pytest.raises(InvalidInputError, match="重複"):
        ExcelComparer().compare(
            old,
            new,
            CompareOptions(algorithm=CompareAlgorithm.KEY_COLUMNS, key_columns=("A", "A")),
        )


def test_encrypted_workbook_signature_has_specific_error(tmp_path: Path) -> None:
    encrypted = tmp_path / "encrypted.xlsx"
    encrypted.write_bytes(bytes.fromhex("D0CF11E0A1B11AE1") + b"encrypted")
    try:
        CompareUseCase().reader.read(encrypted)
    except PasswordProtectedWorkbookError:
        pass
    else:
        raise AssertionError("encrypted workbook was not rejected")


def test_use_case_writes_report_and_highlights(tmp_path: Path) -> None:
    old_path, new_path, output = tmp_path / "old.xlsx", tmp_path / "new.xlsx", tmp_path / "result.xlsx"
    for path, values in [(old_path, {"A1": "old"}), (new_path, {"A1": "new", "B1": "add"})]:
        book = Workbook()
        sheet = book.active
        sheet.title = "Data"
        for coordinate, value in values.items():
            sheet[coordinate] = value
        book.save(path)
    result = CompareUseCase().execute(old_path, new_path, output, CompareOptions())
    assert result.count(DifferenceType.MODIFIED) == 1
    assert result.count(DifferenceType.ADDED) == 1
    book = load_workbook(output)
    assert book.sheetnames[0] == "比較結果"
    assert book["比較結果"]["F11"].hyperlink.target == "#'Data'!A1"
    assert book["Data"]["A1"].fill.fgColor.rgb.endswith("FFF2CC")
    assert book["Data"]["B1"].fill.fgColor.rgb.endswith("C6EFCE")
    book.close()
