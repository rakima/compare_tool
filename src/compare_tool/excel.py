from __future__ import annotations

import os
import shutil
import tempfile
from dataclasses import dataclass
from pathlib import Path
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.worksheet.worksheet import Worksheet

from .comparer import Comparer
from .errors import OutputWriteError, PasswordProtectedWorkbookError, WorkbookReadError
from .models import CompareOptions, CompareResult, Difference, DifferenceType


@dataclass(frozen=True, slots=True)
class CellData:
    value: object = None
    formula: str | None = None

    @property
    def exists(self) -> bool:
        return self.value is not None or self.formula is not None


@dataclass(slots=True)
class ExcelDocument:
    sheets: dict[str, dict[str, CellData]]


class ExcelReader:
    def read(self, path: Path) -> ExcelDocument:
        # Password-protected OOXML is commonly wrapped in an OLE compound file.
        # Detect it before openpyxl reports the less useful "not a zip file".
        try:
            with path.open("rb") as stream:
                if stream.read(8) == bytes.fromhex("D0CF11E0A1B11AE1"):
                    raise PasswordProtectedWorkbookError(f"パスワード付きExcelは比較できません: {path}")
        except PasswordProtectedWorkbookError:
            raise
        except OSError as exc:
            raise WorkbookReadError(f"ファイルを読み取れません: {path}") from exc
        try:
            formulas_book = load_workbook(path, data_only=False, read_only=True)
            values_book = load_workbook(path, data_only=True, read_only=True)
        except PermissionError as exc:
            raise WorkbookReadError(f"ファイルを読み取れません: {path}") from exc
        except (BadZipFile, InvalidFileException, KeyError, OSError, ValueError) as exc:
            message = str(exc).lower()
            if "password" in message or "encrypted" in message:
                raise PasswordProtectedWorkbookError(f"パスワード付きExcelは比較できません: {path}") from exc
            raise WorkbookReadError(f"Excelファイルが破損しているか、読み取れません: {path}") from exc

        try:
            sheets: dict[str, dict[str, CellData]] = {}
            for formula_sheet in formulas_book.worksheets:
                value_sheet = values_book[formula_sheet.title]
                cells: dict[str, CellData] = {}
                for row in formula_sheet.iter_rows():
                    for formula_cell in row:
                        raw = formula_cell.value
                        formula = raw if formula_cell.data_type == "f" else None
                        value = value_sheet[formula_cell.coordinate].value if formula else raw
                        data = CellData(value=value, formula=formula)
                        if data.exists:
                            cells[formula_cell.coordinate] = data
                sheets[formula_sheet.title] = cells
            return ExcelDocument(sheets)
        finally:
            formulas_book.close()
            values_book.close()


class ExcelComparer(Comparer[ExcelDocument]):
    def compare(self, old: ExcelDocument, new: ExcelDocument, options: CompareOptions) -> CompareResult:
        differences: list[Difference] = []
        old_names = set(old.sheets)
        new_names = set(new.sheets)

        for name in sorted(new_names - old_names):
            differences.append(Difference(DifferenceType.SHEET_ADDED, name))
        for name in sorted(old_names - new_names):
            differences.append(Difference(DifferenceType.SHEET_DELETED, name))

        for name in sorted(old_names & new_names):
            differences.extend(self._compare_sheet(name, old.sheets[name], new.sheets[name], options))
        return CompareResult(differences)

    def _compare_sheet(
        self,
        sheet: str,
        old_cells: dict[str, CellData],
        new_cells: dict[str, CellData],
        options: CompareOptions,
    ) -> list[Difference]:
        result: list[Difference] = []
        for coordinate in sorted(set(old_cells) | set(new_cells)):
            old = old_cells.get(coordinate, CellData())
            new = new_cells.get(coordinate, CellData())
            value_changed = options.compare_values and not self._equal(old.value, new.value, options)
            formula_changed = options.compare_formulas and not self._equal(old.formula, new.formula, options)
            if not (value_changed or formula_changed):
                continue

            if not old.exists:
                kind = DifferenceType.ADDED
            elif not new.exists:
                kind = DifferenceType.DELETED
            else:
                kind = DifferenceType.MODIFIED
            old_display = old.formula if formula_changed and old.formula is not None else old.value
            new_display = new.formula if formula_changed and new.formula is not None else new.value
            result.append(Difference(kind, sheet, coordinate, old_display, new_display, value_changed, formula_changed))
        return result

    @staticmethod
    def _equal(left: object, right: object, options: CompareOptions) -> bool:
        def normalize(value: object) -> object:
            if options.empty_string_equals_empty and value == "":
                value = None
            if isinstance(value, str):
                if options.ignore_surrounding_whitespace:
                    value = value.strip()
                if options.ignore_case:
                    value = value.casefold()
            return value

        return normalize(left) == normalize(right)


class ExcelReportWriter:
    REPORT_NAME = "比較結果"
    MODIFIED_FILL = PatternFill("solid", fgColor="FFF2CC")
    ADDED_FILL = PatternFill("solid", fgColor="C6EFCE")
    HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")

    def write(self, source_new: Path, output: Path, result: CompareResult, detailed: bool = True) -> Path:
        workbook = None
        temporary_output: Path | None = None
        try:
            output.parent.mkdir(parents=True, exist_ok=True)
            file_descriptor, temporary_name = tempfile.mkstemp(
                dir=output.parent,
                prefix=f".{output.stem}_",
                suffix=".tmp.xlsx",
            )
            os.close(file_descriptor)
            temporary_output = Path(temporary_name)
            shutil.copy2(source_new, temporary_output)
            workbook = load_workbook(temporary_output)
            report_name = self._unique_report_name(workbook.sheetnames)
            report = workbook.create_sheet(report_name, 0)
            self._write_report(report, result, detailed)
            self._highlight_differences(workbook, result)
            workbook.save(temporary_output)
            # Close before replacing because Windows does not allow replacing
            # a file that is still opened by openpyxl/zipfile.
            workbook.close()
            workbook = None
            os.replace(temporary_output, output)
            temporary_output = None
            result.output_path = output
            return output
        except (PermissionError, OSError, InvalidFileException, BadZipFile, ValueError, shutil.Error) as exc:
            raise OutputWriteError(f"出力ファイルを保存できません: {output}") from exc
        finally:
            if workbook is not None:
                workbook.close()
            if temporary_output is not None:
                try:
                    temporary_output.unlink(missing_ok=True)
                except OSError:
                    pass

    def _write_report(self, sheet: Worksheet, result: CompareResult, detailed: bool) -> None:
        sheet["A1"] = "比較サマリー"
        sheet["A1"].font = Font(bold=True, size=14)
        labels = [
            (DifferenceType.MODIFIED, "変更件数"),
            (DifferenceType.ADDED, "追加件数"),
            (DifferenceType.DELETED, "削除件数"),
            (DifferenceType.SHEET_ADDED, "シート追加件数"),
            (DifferenceType.SHEET_DELETED, "シート削除件数"),
        ]
        for row, (kind, label) in enumerate(labels, 2):
            sheet.cell(row, 1, label)
            sheet.cell(row, 2, result.count(kind))

        if not detailed:
            self._finish_layout(sheet)
            return

        header_row = 8
        headers = ["種別", "シート", "セル", "旧値", "新値", "リンク"]
        for column, value in enumerate(headers, 1):
            cell = sheet.cell(header_row, column, value)
            cell.font = Font(bold=True)
            cell.fill = self.HEADER_FILL
        for row, difference in enumerate(result.differences, header_row + 1):
            values = [difference.kind.value, difference.sheet, difference.cell or "", difference.old_value, difference.new_value]
            for column, value in enumerate(values, 1):
                sheet.cell(row, column, self._display(value))
            if difference.can_link:
                link = sheet.cell(row, 6, "ジャンプ")
                escaped_sheet = difference.sheet.replace("'", "''")
                link.hyperlink = f"#'{escaped_sheet}'!{difference.cell}"
                link.style = "Hyperlink"
        sheet.auto_filter.ref = f"A{header_row}:F{max(header_row, header_row + result.total)}"
        sheet.freeze_panes = f"A{header_row + 1}"
        self._finish_layout(sheet)

    @staticmethod
    def _display(value: object) -> object:
        if value is None:
            return ""
        text = str(value)
        return f"'{text}" if text.startswith("=") else value

    @staticmethod
    def _finish_layout(sheet: Worksheet) -> None:
        widths = {"A": 18, "B": 24, "C": 12, "D": 36, "E": 36, "F": 12}
        for column, width in widths.items():
            sheet.column_dimensions[column].width = width
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    def _highlight_differences(self, workbook, result: CompareResult) -> None:
        for difference in result.differences:
            if not difference.cell or difference.sheet not in workbook.sheetnames:
                continue
            if difference.kind is DifferenceType.MODIFIED:
                workbook[difference.sheet][difference.cell].fill = self.MODIFIED_FILL
            elif difference.kind is DifferenceType.ADDED:
                workbook[difference.sheet][difference.cell].fill = self.ADDED_FILL

    def _unique_report_name(self, names: list[str]) -> str:
        if self.REPORT_NAME not in names:
            return self.REPORT_NAME
        index = 1
        while f"{self.REPORT_NAME}_{index}" in names:
            index += 1
        return f"{self.REPORT_NAME}_{index}"
