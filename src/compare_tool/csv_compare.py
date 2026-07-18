from __future__ import annotations

import csv
import os
import tempfile
from contextlib import suppress
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.worksheet.worksheet import Worksheet

from .comparer import CancelCheck, Comparer
from .errors import OutputWriteError, WorkbookReadError
from .excel import CellData, ExcelComparer, ExcelDocument, ExcelReportWriter, raise_if_cancelled
from .models import CompareOptions, CompareResult

CSV_SHEET_NAME = "CSV"
CSV_AUTO_ENCODING = "auto"
CSV_AUTO_DELIMITER = "auto"
CSV_DELIMITER_CANDIDATES = (",", "\t", ";")
CSV_ENCODING_LABELS = {
    "auto": "自動",
    "utf-8-sig": "UTF-8 / UTF-8 BOM",
    "utf-8": "UTF-8",
    "cp932": "Shift_JIS",
}
CSV_DELIMITER_LABELS = {
    "auto": "自動",
    ",": "カンマ",
    "\t": "タブ",
    ";": "セミコロン",
}


@dataclass(slots=True)
class CsvDocument:
    rows: list[list[str]]

    @property
    def cells(self) -> dict[str, CellData]:
        cells: dict[str, CellData] = {}
        for row_index, row in enumerate(self.rows, 1):
            for column_index, value in enumerate(row, 1):
                if value != "":
                    cells[f"{get_column_letter(column_index)}{row_index}"] = CellData(value=value)
        return cells


class CsvReader:
    def read(self, path: Path, options: CompareOptions | None = None) -> CsvDocument:
        options = options or CompareOptions()
        encoding = self._resolve_encoding(path, options.csv_encoding)
        delimiter = self._resolve_delimiter(path, encoding, options.csv_delimiter)
        try:
            with path.open("r", encoding=encoding, newline="") as stream:
                rows = [
                    row
                    for row in csv.reader(stream, delimiter=delimiter)
                    if not (options.ignore_csv_blank_lines and self._is_blank_row(row))
                ]
                return CsvDocument(rows)
        except LookupError as exc:
            raise WorkbookReadError(
                f"CSVの文字コード指定が不正です: {options.csv_encoding}\n"
                "文字コードは「自動」「UTF-8 / UTF-8 BOM」「Shift_JIS」から選択してください。"
            ) from exc
        except UnicodeDecodeError as exc:
            raise WorkbookReadError(
                f"CSVファイルを {encoding} として読み取れません: {path}\n"
                "CSV文字コードの指定が実際のファイルと異なる可能性があります。"
                "「自動」で失敗する場合は「UTF-8 / UTF-8 BOM」または「Shift_JIS」を選び直してください。"
            ) from exc
        except OSError as exc:
            raise WorkbookReadError(
                f"CSVファイルを読み取れません: {path}\n"
                "ファイルが存在するか、他のアプリで使用中ではないか、読み取り権限があるか確認してください。"
            ) from exc
        except csv.Error as exc:
            raise WorkbookReadError(
                f"CSVファイルの形式を読み取れません: {path}\n"
                "引用符や改行を含むフィールドが壊れている可能性があります。"
                "Excelなどで開けるか確認し、必要ならCSVとして保存し直してください。"
            ) from exc

    def _resolve_encoding(self, path: Path, selected_encoding: str) -> str:
        if selected_encoding != CSV_AUTO_ENCODING:
            return selected_encoding
        data = self._read_sample(path)
        if data.startswith(b"\xef\xbb\xbf"):
            return "utf-8-sig"
        if data.startswith((b"\xff\xfe", b"\xfe\xff")):
            raise self._auto_detection_error(path)
        for encoding in ("utf-8", "cp932"):
            try:
                text = data.decode(encoding)
            except UnicodeDecodeError:
                continue
            if self._looks_like_text(text):
                return encoding
            continue
        raise self._auto_detection_error(path)

    @staticmethod
    def _looks_like_text(text: str) -> bool:
        return not any((ord(character) < 32 and character not in "\r\n\t") for character in text)

    @staticmethod
    def _is_blank_row(row: list[str]) -> bool:
        return all(value == "" for value in row)

    @staticmethod
    def _auto_detection_error(path: Path) -> WorkbookReadError:
        return WorkbookReadError(
            f"CSVファイルの文字コードを自動判定できません: {path}\n"
            "対応している自動判定候補はUTF-8 / UTF-8 BOM / Shift_JISです。"
            "別の文字コードのCSVはUTF-8またはShift_JISで保存し直してください。"
        )

    def _resolve_delimiter(self, path: Path, encoding: str, selected_delimiter: str) -> str:
        if selected_delimiter != CSV_AUTO_DELIMITER:
            return selected_delimiter
        rows = self._read_rows_for_detection(path, encoding)
        scores = [(self._delimiter_score(rows, delimiter), delimiter) for delimiter in CSV_DELIMITER_CANDIDATES]
        best_score, best_delimiter = max(scores, key=lambda item: item[0])
        return "," if best_score <= 0 else best_delimiter

    @staticmethod
    def _delimiter_score(lines: list[str], delimiter: str) -> float:
        parsed_rows = list(csv.reader(lines, delimiter=delimiter))
        column_counts = [len(row) for row in parsed_rows if row]
        if not column_counts:
            return 0
        max_columns = max(column_counts)
        if max_columns <= 1:
            return 0
        most_common_columns = max(set(column_counts), key=column_counts.count)
        consistency = column_counts.count(most_common_columns) / len(column_counts)
        average_columns = sum(column_counts) / len(column_counts)
        return consistency * 100 + average_columns * 10

    @staticmethod
    def _read_rows_for_detection(path: Path, encoding: str, limit: int = 20) -> list[str]:
        try:
            with path.open("r", encoding=encoding, newline="") as stream:
                rows: list[str] = []
                for _ in range(limit):
                    line = stream.readline()
                    if line == "":
                        break
                    rows.append(line)
                return rows
        except UnicodeDecodeError as exc:
            raise WorkbookReadError(
                f"CSVファイルを {encoding} として読み取れません: {path}\n"
                "CSV文字コードの指定が実際のファイルと異なる可能性があります。"
                "「自動」で失敗する場合は「UTF-8 / UTF-8 BOM」または「Shift_JIS」を選び直してください。"
            ) from exc
        except OSError as exc:
            raise WorkbookReadError(
                f"CSVファイルを読み取れません: {path}\n"
                "ファイルが存在するか、他のアプリで使用中ではないか、読み取り権限があるか確認してください。"
            ) from exc

    @staticmethod
    def _read_sample(path: Path) -> bytes:
        try:
            with path.open("rb") as stream:
                return stream.read()
        except OSError as exc:
            raise WorkbookReadError(
                f"CSVファイルを読み取れません: {path}\n"
                "ファイルが存在するか、他のアプリで使用中ではないか、読み取り権限があるか確認してください。"
            ) from exc


class CsvComparer(Comparer[CsvDocument]):
    def __init__(self, table_comparer: ExcelComparer | None = None) -> None:
        self.table_comparer = table_comparer or ExcelComparer()

    def compare(
        self,
        old: CsvDocument,
        new: CsvDocument,
        options: CompareOptions,
        cancel_requested: CancelCheck | None = None,
    ) -> CompareResult:
        return self.table_comparer.compare(
            ExcelDocument({CSV_SHEET_NAME: old.cells}),
            ExcelDocument({CSV_SHEET_NAME: new.cells}),
            options,
            cancel_requested,
        )


class CsvReportWriter(ExcelReportWriter):
    def write(
        self,
        source_new: Path,
        output: Path,
        result: CompareResult,
        detailed: bool = True,
        cancel_requested: CancelCheck | None = None,
        options: CompareOptions | None = None,
    ) -> Path:
        workbook = None
        temporary_output: Path | None = None
        try:
            self._raise_if_cancelled(cancel_requested)
            output.parent.mkdir(parents=True, exist_ok=True)
            file_descriptor, temporary_name = tempfile.mkstemp(
                dir=output.parent,
                prefix=f".{output.stem}_",
                suffix=".tmp.xlsx",
            )
            os.close(file_descriptor)
            temporary_output = Path(temporary_name)

            csv_document = CsvReader().read(source_new, options)
            options = options or CompareOptions()
            workbook = Workbook()
            csv_sheet = workbook.active
            csv_sheet.title = CSV_SHEET_NAME
            self._write_csv_sheet(csv_sheet, csv_document, cancel_requested)
            report = workbook.create_sheet(self._unique_report_name(workbook.sheetnames), 0)
            self._write_report(report, result, detailed, cancel_requested)
            self._write_csv_settings(report, options)
            self._highlight_differences(workbook, result, cancel_requested)
            self._raise_if_cancelled(cancel_requested)
            workbook.save(temporary_output)
            workbook.close()
            workbook = None
            os.replace(temporary_output, output)
            temporary_output = None
            result.output_path = output
            return output
        except (PermissionError, OSError, InvalidFileException, ValueError, csv.Error) as exc:
            raise OutputWriteError(f"出力ファイルを保存できません: {output}") from exc
        finally:
            if workbook is not None:
                workbook.close()
            if temporary_output is not None:
                with suppress(OSError):
                    temporary_output.unlink(missing_ok=True)

    @staticmethod
    def _write_csv_sheet(
        sheet: Worksheet,
        document: CsvDocument,
        cancel_requested: CancelCheck | None = None,
    ) -> None:
        for row_index, row in enumerate(document.rows, 1):
            raise_if_cancelled(cancel_requested)
            for column_index, value in enumerate(row, 1):
                sheet.cell(row_index, column_index, value)

    @classmethod
    def _write_csv_settings(cls, sheet: Worksheet, options: CompareOptions) -> None:
        sheet["H1"] = "CSV読み込み設定"
        sheet["H1"].font = Font(bold=True, size=14)
        rows = [
            ("文字コード", CSV_ENCODING_LABELS.get(options.csv_encoding, options.csv_encoding)),
            ("区切り文字", CSV_DELIMITER_LABELS.get(options.csv_delimiter, options.csv_delimiter)),
            ("空行を無視", "はい" if options.ignore_csv_blank_lines else "いいえ"),
        ]
        for row_index, (label, value) in enumerate(rows, 2):
            cls._write_cell(sheet, row_index, 8, label)
            cls._write_cell(sheet, row_index, 9, value)
        sheet.column_dimensions["H"].width = 18
        sheet.column_dimensions["I"].width = 24
