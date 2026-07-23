from __future__ import annotations

import json
import os
import tempfile
from contextlib import suppress
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.worksheet.worksheet import Worksheet

from .comparer import CancelCheck, Comparer
from .errors import InvalidInputError, OperationCancelledError, OutputWriteError, WorkbookReadError
from .excel import ExcelReportWriter
from .models import CompareOptions, CompareResult, Difference, DifferenceType

JSON_SHEET_NAME = "JSON"


@dataclass(frozen=True, slots=True)
class JsonDocument:
    data: Any


class JsonReader:
    def read(self, path: Path) -> JsonDocument:
        try:
            with path.open("r", encoding="utf-8-sig") as stream:
                return JsonDocument(json.load(stream))
        except UnicodeDecodeError as exc:
            raise WorkbookReadError(
                f"JSONファイルをUTF-8として読み取れません: {path}\n"
                "JSONはUTF-8 / UTF-8 BOM付きで保存してください。"
                "文字化けする場合は、エディタでUTF-8として保存し直してから再実行してください。"
            ) from exc
        except json.JSONDecodeError as exc:
            raise WorkbookReadError(
                f"JSONファイルの形式を読み取れません: {path}\n"
                f"{exc.lineno}行 {exc.colno}列付近でJSON構文エラーが発生しました: {exc.msg}\n"
                "カンマ、引用符、コロン、括弧の対応を確認してください。"
            ) from exc
        except OSError as exc:
            raise WorkbookReadError(
                f"JSONファイルを読み取れません: {path}\n"
                "ファイルが存在するか、他のアプリで使用中ではないか、読み取り権限があるか確認してください。"
            ) from exc


class JsonComparer(Comparer[JsonDocument]):
    def compare(
        self,
        old: JsonDocument,
        new: JsonDocument,
        options: CompareOptions,
        cancel_requested: CancelCheck | None = None,
    ) -> CompareResult:
        differences: list[Difference] = []
        self._compare_value("$", old.data, new.data, options, differences, cancel_requested)
        return CompareResult(differences)

    def _compare_value(
        self,
        path: str,
        old: Any,
        new: Any,
        options: CompareOptions,
        differences: list[Difference],
        cancel_requested: CancelCheck | None,
    ) -> None:
        self._raise_if_cancelled(cancel_requested)
        if isinstance(old, dict) and isinstance(new, dict):
            self._compare_object(path, old, new, options, differences, cancel_requested)
            return
        if isinstance(old, list) and isinstance(new, list):
            self._compare_array(path, old, new, options, differences, cancel_requested)
            return
        if not self._equal(old, new, options):
            differences.append(
                Difference(
                    DifferenceType.MODIFIED,
                    JSON_SHEET_NAME,
                    path,
                    old,
                    new,
                    value_changed=True,
                )
            )

    def _compare_object(
        self,
        path: str,
        old: dict[str, Any],
        new: dict[str, Any],
        options: CompareOptions,
        differences: list[Difference],
        cancel_requested: CancelCheck | None,
    ) -> None:
        if not options.ignore_json_object_key_order and list(old.keys()) != list(new.keys()):
            differences.append(
                Difference(
                    DifferenceType.MODIFIED,
                    JSON_SHEET_NAME,
                    path,
                    list(old.keys()),
                    list(new.keys()),
                    value_changed=True,
                )
            )
        for key in sorted(old.keys() - new.keys()):
            self._raise_if_cancelled(cancel_requested)
            differences.append(
                Difference(
                    DifferenceType.DELETED,
                    JSON_SHEET_NAME,
                    self._object_path(path, key),
                    old[key],
                )
            )
        for key in sorted(new.keys() - old.keys()):
            self._raise_if_cancelled(cancel_requested)
            differences.append(
                Difference(DifferenceType.ADDED, JSON_SHEET_NAME, self._object_path(path, key), None, new[key])
            )
        for key in sorted(old.keys() & new.keys()):
            self._compare_value(
                self._object_path(path, key),
                old[key],
                new[key],
                options,
                differences,
                cancel_requested,
            )

    def _compare_array(
        self,
        path: str,
        old: list[Any],
        new: list[Any],
        options: CompareOptions,
        differences: list[Difference],
        cancel_requested: CancelCheck | None,
    ) -> None:
        if options.json_array_key and self._can_compare_keyed_array(old, new, options.json_array_key):
            self._compare_keyed_array(path, old, new, options, differences, cancel_requested)
            return

        if options.ignore_json_array_order:
            self._compare_unordered_array(path, old, new, options, differences, cancel_requested)
            return

        common_length = min(len(old), len(new))
        for index in range(common_length):
            self._compare_value(f"{path}[{index}]", old[index], new[index], options, differences, cancel_requested)
        for index in range(common_length, len(old)):
            self._raise_if_cancelled(cancel_requested)
            differences.append(Difference(DifferenceType.DELETED, JSON_SHEET_NAME, f"{path}[{index}]", old[index]))
        for index in range(common_length, len(new)):
            self._raise_if_cancelled(cancel_requested)
            differences.append(Difference(DifferenceType.ADDED, JSON_SHEET_NAME, f"{path}[{index}]", None, new[index]))

    @staticmethod
    def _object_path(base: str, key: str) -> str:
        if key.isidentifier():
            return f"{base}.{key}"
        escaped = key.replace("\\", "\\\\").replace("'", "\\'")
        return f"{base}['{escaped}']"

    @staticmethod
    def _equal(left: Any, right: Any, options: CompareOptions) -> bool:
        def normalize(value: Any) -> Any:
            if options.empty_string_equals_empty and value == "":
                value = None
            if isinstance(value, str):
                if options.ignore_surrounding_whitespace:
                    value = value.strip()
                if options.ignore_case:
                    value = value.casefold()
            return value

        return normalize(left) == normalize(right)

    @classmethod
    def _compare_unordered_array(
        cls,
        path: str,
        old: list[Any],
        new: list[Any],
        options: CompareOptions,
        differences: list[Difference],
        cancel_requested: CancelCheck | None,
    ) -> None:
        unmatched_new = {index: cls._canonical_json_value(value, options) for index, value in enumerate(new)}
        unmatched_old: list[int] = []

        for old_index, old_value in enumerate(old):
            cls._raise_if_cancelled(cancel_requested)
            canonical_old = cls._canonical_json_value(old_value, options)
            new_index = cls._find_first_match(canonical_old, unmatched_new)
            if new_index is None:
                unmatched_old.append(old_index)
            else:
                del unmatched_new[new_index]

        for old_index in unmatched_old:
            cls._raise_if_cancelled(cancel_requested)
            differences.append(
                Difference(DifferenceType.DELETED, JSON_SHEET_NAME, f"{path}[{old_index}]", old[old_index])
            )
        for new_index in sorted(unmatched_new):
            cls._raise_if_cancelled(cancel_requested)
            differences.append(
                Difference(DifferenceType.ADDED, JSON_SHEET_NAME, f"{path}[{new_index}]", None, new[new_index])
            )

    @staticmethod
    def _find_first_match(value: str, candidates: dict[int, str]) -> int | None:
        for index, candidate in candidates.items():
            if candidate == value:
                return index
        return None

    def _compare_keyed_array(
        self,
        path: str,
        old: list[Any],
        new: list[Any],
        options: CompareOptions,
        differences: list[Difference],
        cancel_requested: CancelCheck | None,
    ) -> None:
        key = options.json_array_key.strip()
        old_items = self._keyed_items(path, old, key, options, "旧JSON")
        new_items = self._keyed_items(path, new, key, options, "新JSON")

        old_keys = set(old_items)
        new_keys = set(new_items)
        for item_key in sorted(old_keys - new_keys, key=self._sort_key):
            self._raise_if_cancelled(cancel_requested)
            differences.append(
                Difference(
                    DifferenceType.DELETED,
                    JSON_SHEET_NAME,
                    self._keyed_array_path(path, key, old_items[item_key][0]),
                    old_items[item_key][1],
                )
            )
        for item_key in sorted(new_keys - old_keys, key=self._sort_key):
            self._raise_if_cancelled(cancel_requested)
            differences.append(
                Difference(
                    DifferenceType.ADDED,
                    JSON_SHEET_NAME,
                    self._keyed_array_path(path, key, new_items[item_key][0]),
                    None,
                    new_items[item_key][1],
                )
            )
        for item_key in sorted(old_keys & new_keys, key=self._sort_key):
            self._compare_value(
                self._keyed_array_path(path, key, new_items[item_key][0]),
                old_items[item_key][1],
                new_items[item_key][1],
                options,
                differences,
                cancel_requested,
            )

    @classmethod
    def _keyed_items(
        cls,
        path: str,
        items: list[Any],
        key: str,
        options: CompareOptions,
        label: str,
    ) -> dict[str, tuple[Any, dict[str, Any]]]:
        keyed: dict[str, tuple[Any, dict[str, Any]]] = {}
        for index, item in enumerate(items):
            raw_key = item[key]
            normalized_key = cls._canonical_json_value(raw_key, options)
            if normalized_key in keyed:
                raise InvalidInputError(
                    f"JSON配列キー `{key}` の値が重複しています: {path}[{index}] ({label})\n"
                    "キー指定比較では、同じ配列内のキー値が一意になるようにしてください。"
                )
            keyed[normalized_key] = (raw_key, item)
        return keyed

    @staticmethod
    def _can_compare_keyed_array(old: list[Any], new: list[Any], key: str) -> bool:
        normalized_key = key.strip()
        if not normalized_key:
            return False
        return all(isinstance(item, dict) and normalized_key in item for item in [*old, *new])

    @staticmethod
    def _keyed_array_path(base: str, key: str, value: Any) -> str:
        display_value = json.dumps(value, ensure_ascii=False, sort_keys=True)
        if key.isidentifier():
            return f"{base}[{key}={display_value}]"
        escaped = key.replace("\\", "\\\\").replace("'", "\\'")
        return f"{base}['{escaped}'={display_value}]"

    @staticmethod
    def _sort_key(value: str) -> str:
        return value

    @classmethod
    def _canonical_json_value(cls, value: Any, options: CompareOptions) -> str:
        return json.dumps(
            cls._normalize_json_value(value, options),
            ensure_ascii=False,
            sort_keys=True,
            separators=(",", ":"),
        )

    @classmethod
    def _normalize_json_value(cls, value: Any, options: CompareOptions) -> Any:
        if options.empty_string_equals_empty and value == "":
            return None
        if isinstance(value, str):
            if options.ignore_surrounding_whitespace:
                value = value.strip()
            if options.ignore_case:
                value = value.casefold()
            return value
        if isinstance(value, dict):
            return {key: cls._normalize_json_value(item, options) for key, item in value.items()}
        if isinstance(value, list):
            return [cls._normalize_json_value(item, options) for item in value]
        return value

    @staticmethod
    def _raise_if_cancelled(cancel_requested: CancelCheck | None) -> None:
        if cancel_requested is not None and cancel_requested():
            raise OperationCancelledError("比較をキャンセルしました。")


class JsonReportWriter(ExcelReportWriter):
    def write(
        self,
        source_new: Path,
        output: Path,
        result: CompareResult,
        detailed: bool = True,
        cancel_requested: CancelCheck | None = None,
        options: CompareOptions | None = None,
    ) -> Path:
        workbook = None
        temporary_output: Path | None = None
        try:
            self._raise_if_cancelled(cancel_requested)
            output.parent.mkdir(parents=True, exist_ok=True)
            file_descriptor, temporary_name = tempfile.mkstemp(
                dir=output.parent,
                prefix=f".{output.stem}_",
                suffix=".tmp.xlsx",
            )
            os.close(file_descriptor)
            temporary_output = Path(temporary_name)

            json_document = JsonReader().read(source_new)
            workbook = Workbook()
            json_sheet = workbook.active
            json_sheet.title = JSON_SHEET_NAME
            self._write_json_sheet(json_sheet, json_document, cancel_requested)
            report = workbook.create_sheet(self._unique_report_name(workbook.sheetnames), 0)
            self._write_report(report, self._displayable_result(result), detailed, cancel_requested)
            self._remove_json_links(report, result)
            self._write_json_settings(report, options or CompareOptions())
            self._raise_if_cancelled(cancel_requested)
            workbook.save(temporary_output)
            workbook.close()
            workbook = None
            os.replace(temporary_output, output)
            temporary_output = None
            result.output_path = output
            return output
        except (PermissionError, OSError, InvalidFileException, ValueError, TypeError) as exc:
            raise OutputWriteError(f"出力ファイルを保存できません: {output}") from exc
        finally:
            if workbook is not None:
                workbook.close()
            if temporary_output is not None:
                with suppress(OSError):
                    temporary_output.unlink(missing_ok=True)

    def _write_json_sheet(
        self,
        sheet: Worksheet,
        document: JsonDocument,
        cancel_requested: CancelCheck | None = None,
    ) -> None:
        sheet["A1"] = "新JSON"
        sheet["A1"].font = Font(bold=True, size=14)
        text = json.dumps(document.data, ensure_ascii=False, indent=2)
        for row_index, line in enumerate(text.splitlines(), 2):
            self._raise_if_cancelled(cancel_requested)
            self._write_cell(sheet, row_index, 1, line)
        sheet.column_dimensions["A"].width = 100

    @staticmethod
    def _displayable_result(result: CompareResult) -> CompareResult:
        return CompareResult(
            [
                Difference(
                    difference.kind,
                    difference.sheet,
                    difference.cell,
                    JsonReportWriter._display_json_value(difference.old_value),
                    JsonReportWriter._display_json_value(difference.new_value),
                    difference.value_changed,
                    difference.formula_changed,
                )
                for difference in result.differences
            ],
            result.output_path,
        )

    @staticmethod
    def _display_json_value(value: object) -> object:
        if isinstance(value, (dict, list)):
            return json.dumps(value, ensure_ascii=False)
        return value

    @staticmethod
    def _remove_json_links(sheet: Worksheet, result: CompareResult) -> None:
        for row_index in range(11, 11 + result.total):
            cell = sheet.cell(row_index, 6)
            cell.value = None
            cell.hyperlink = None
            cell.style = "Normal"

    @classmethod
    def _write_json_settings(cls, sheet: Worksheet, options: CompareOptions) -> None:
        sheet["H1"] = "JSON読み込み設定"
        sheet["H1"].font = Font(bold=True, size=14)
        rows = [
            ("文字コード", "UTF-8 / UTF-8 BOM"),
            ("比較位置", "JSON Path"),
            ("オブジェクトのキー順を無視", "はい" if options.ignore_json_object_key_order else "いいえ"),
            ("配列順序を無視", "はい" if options.ignore_json_array_order else "いいえ"),
            ("配列キー", options.json_array_key or "未指定"),
        ]
        for row_index, (label, value) in enumerate(rows, 2):
            cls._write_cell(sheet, row_index, 8, label)
            cls._write_cell(sheet, row_index, 9, value)
        sheet.column_dimensions["H"].width = 18
        sheet.column_dimensions["I"].width = 24
