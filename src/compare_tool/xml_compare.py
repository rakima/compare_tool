from __future__ import annotations

import os
import tempfile
from contextlib import suppress
from dataclasses import dataclass
from pathlib import Path
from typing import Any
from xml.etree import ElementTree

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.worksheet.worksheet import Worksheet

from .comparer import CancelCheck, Comparer
from .errors import OperationCancelledError, OutputWriteError, WorkbookReadError
from .excel import ExcelReportWriter
from .models import CompareOptions, CompareResult, Difference, DifferenceType

XML_SHEET_NAME = "XML"


@dataclass(frozen=True, slots=True)
class XmlDocument:
    root: ElementTree.Element


class XmlReader:
    def read(self, path: Path) -> XmlDocument:
        try:
            parser = ElementTree.XMLParser(target=ElementTree.TreeBuilder(insert_comments=False))
            with path.open("r", encoding="utf-8-sig") as stream:
                return XmlDocument(ElementTree.parse(stream, parser=parser).getroot())
        except UnicodeDecodeError as exc:
            raise WorkbookReadError(
                f"XMLファイルをUTF-8として読み取れません: {path}\n"
                "XMLはUTF-8 / UTF-8 BOM付きで保存してください。"
                "文字化けする場合は、エディタでUTF-8として保存し直してから再実行してください。"
            ) from exc
        except ElementTree.ParseError as exc:
            raise WorkbookReadError(
                f"XMLファイルの形式を読み取れません: {path}\n"
                f"{exc.position[0]}行 {exc.position[1]}列付近でXML構文エラーが発生しました: {exc}。\n"
                "開始タグと終了タグの対応、属性の引用符、特殊文字のエスケープを確認してください。"
            ) from exc
        except OSError as exc:
            raise WorkbookReadError(
                f"XMLファイルを読み取れません: {path}\n"
                "ファイルが存在するか、他のアプリで使用中ではないか、読み取り権限があるか確認してください。"
            ) from exc


class XmlComparer(Comparer[XmlDocument]):
    def compare(
        self,
        old: XmlDocument,
        new: XmlDocument,
        options: CompareOptions,
        cancel_requested: CancelCheck | None = None,
    ) -> CompareResult:
        differences: list[Difference] = []
        self._compare_element(
            f"/{self._display_tag(old.root.tag)}",
            old.root,
            new.root,
            options,
            differences,
            cancel_requested,
        )
        return CompareResult(differences)

    def _compare_element(
        self,
        path: str,
        old: ElementTree.Element,
        new: ElementTree.Element,
        options: CompareOptions,
        differences: list[Difference],
        cancel_requested: CancelCheck | None,
    ) -> None:
        self._raise_if_cancelled(cancel_requested)
        if old.tag != new.tag:
            differences.append(
                Difference(
                    DifferenceType.MODIFIED,
                    XML_SHEET_NAME,
                    path,
                    self._display_tag(old.tag),
                    self._display_tag(new.tag),
                    value_changed=True,
                )
            )
            return

        self._compare_attributes(path, old, new, options, differences, cancel_requested)
        self._compare_text(path, old, new, options, differences)
        self._compare_children(path, old, new, options, differences, cancel_requested)

    def _compare_attributes(
        self,
        path: str,
        old: ElementTree.Element,
        new: ElementTree.Element,
        options: CompareOptions,
        differences: list[Difference],
        cancel_requested: CancelCheck | None,
    ) -> None:
        if not options.ignore_xml_attribute_order and list(old.attrib) != list(new.attrib):
            differences.append(
                Difference(
                    DifferenceType.MODIFIED,
                    XML_SHEET_NAME,
                    f"{path}/@*",
                    list(old.attrib),
                    list(new.attrib),
                    value_changed=True,
                )
            )

        for name in sorted(old.attrib.keys() - new.attrib.keys()):
            self._raise_if_cancelled(cancel_requested)
            differences.append(Difference(DifferenceType.DELETED, XML_SHEET_NAME, f"{path}/@{name}", old.attrib[name]))
        for name in sorted(new.attrib.keys() - old.attrib.keys()):
            self._raise_if_cancelled(cancel_requested)
            differences.append(
                Difference(DifferenceType.ADDED, XML_SHEET_NAME, f"{path}/@{name}", None, new.attrib[name])
            )
        for name in sorted(old.attrib.keys() & new.attrib.keys()):
            self._raise_if_cancelled(cancel_requested)
            if not self._equal(old.attrib[name], new.attrib[name], options):
                differences.append(
                    Difference(
                        DifferenceType.MODIFIED,
                        XML_SHEET_NAME,
                        f"{path}/@{name}",
                        old.attrib[name],
                        new.attrib[name],
                        value_changed=True,
                    )
                )

    def _compare_text(
        self,
        path: str,
        old: ElementTree.Element,
        new: ElementTree.Element,
        options: CompareOptions,
        differences: list[Difference],
    ) -> None:
        old_text = self._text_value(old.text, options)
        new_text = self._text_value(new.text, options)
        if not self._equal(old_text, new_text, options):
            differences.append(
                Difference(
                    DifferenceType.MODIFIED,
                    XML_SHEET_NAME,
                    path,
                    old_text,
                    new_text,
                    value_changed=True,
                )
            )

    def _compare_children(
        self,
        path: str,
        old: ElementTree.Element,
        new: ElementTree.Element,
        options: CompareOptions,
        differences: list[Difference],
        cancel_requested: CancelCheck | None,
    ) -> None:
        old_children = list(old)
        new_children = list(new)
        old_child_paths = self._child_paths(path, old_children)
        new_child_paths = self._child_paths(path, new_children)
        keyed_matches = self._keyed_child_matches(old_children, new_children, options)
        matched_old = {old_index for old_index, _new_index in keyed_matches}
        matched_new = {new_index for _old_index, new_index in keyed_matches}

        for old_index, new_index in keyed_matches:
            self._compare_element(
                old_child_paths[old_index],
                old_children[old_index],
                new_children[new_index],
                options,
                differences,
                cancel_requested,
            )

        remaining_old = [child for index, child in enumerate(old_children) if index not in matched_old]
        remaining_new = [child for index, child in enumerate(new_children) if index not in matched_new]
        remaining_old_paths = [
            child_path for index, child_path in enumerate(old_child_paths) if index not in matched_old
        ]
        remaining_new_paths = [
            child_path for index, child_path in enumerate(new_child_paths) if index not in matched_new
        ]

        self._compare_children_by_lcs(
            remaining_old,
            remaining_new,
            remaining_old_paths,
            remaining_new_paths,
            options,
            differences,
            cancel_requested,
        )

    def _compare_children_by_lcs(
        self,
        old_children: list[ElementTree.Element],
        new_children: list[ElementTree.Element],
        old_child_paths: list[str],
        new_child_paths: list[str],
        options: CompareOptions,
        differences: list[Difference],
        cancel_requested: CancelCheck | None,
    ) -> None:
        matches = self._child_lcs_matches(old_children, new_children, options, cancel_requested)
        old_start = 0
        new_start = 0

        for old_index, new_index in matches:
            self._compare_child_gap(
                old_children[old_start:old_index],
                new_children[new_start:new_index],
                old_child_paths[old_start:old_index],
                new_child_paths[new_start:new_index],
                options,
                differences,
                cancel_requested,
            )
            self._compare_element(
                old_child_paths[old_index],
                old_children[old_index],
                new_children[new_index],
                options,
                differences,
                cancel_requested,
            )
            old_start = old_index + 1
            new_start = new_index + 1

        self._compare_child_gap(
            old_children[old_start:],
            new_children[new_start:],
            old_child_paths[old_start:],
            new_child_paths[new_start:],
            options,
            differences,
            cancel_requested,
        )

    @classmethod
    def _keyed_child_matches(
        cls,
        old_children: list[ElementTree.Element],
        new_children: list[ElementTree.Element],
        options: CompareOptions,
    ) -> list[tuple[int, int]]:
        old_keys = cls._unique_child_key_indexes(old_children, options)
        new_keys = cls._unique_child_key_indexes(new_children, options)
        return sorted((old_index, new_keys[key]) for key, old_index in old_keys.items() if key in new_keys)

    @classmethod
    def _unique_child_key_indexes(
        cls,
        children: list[ElementTree.Element],
        options: CompareOptions,
    ) -> dict[tuple[str, str, object], int]:
        indexes: dict[tuple[str, str, object], int] = {}
        duplicated: set[tuple[str, str, object]] = set()
        for index, child in enumerate(children):
            key = cls._child_key(child, options)
            if key is None:
                continue
            if key in indexes:
                duplicated.add(key)
                del indexes[key]
                continue
            if key not in duplicated:
                indexes[key] = index
        return indexes

    @classmethod
    def _child_key(
        cls,
        child: ElementTree.Element,
        options: CompareOptions,
    ) -> tuple[str, str, object] | None:
        for attribute in ("id", "name"):
            if attribute in child.attrib:
                return (
                    cls._display_tag(child.tag),
                    attribute,
                    cls._normalize_value(child.attrib[attribute], options),
                )
        return None

    def _compare_child_gap(
        self,
        old_children: list[ElementTree.Element],
        new_children: list[ElementTree.Element],
        old_child_paths: list[str],
        new_child_paths: list[str],
        options: CompareOptions,
        differences: list[Difference],
        cancel_requested: CancelCheck | None,
    ) -> None:
        index = 0
        while index < len(old_children) and index < len(new_children):
            if old_children[index].tag != new_children[index].tag:
                break
            self._compare_element(
                old_child_paths[index],
                old_children[index],
                new_children[index],
                options,
                differences,
                cancel_requested,
            )
            index += 1

        if index == len(old_children) and index == len(new_children):
            return

        if len(old_children) - index == len(new_children) - index and self._same_tag_sequence(
            old_children[index:],
            new_children[index:],
        ):
            for offset, (old_child, new_child) in enumerate(
                zip(old_children[index:], new_children[index:], strict=True),
                index,
            ):
                self._compare_element(
                    old_child_paths[offset],
                    old_child,
                    new_child,
                    options,
                    differences,
                    cancel_requested,
                )
            return

        remaining_start = index
        for index, child in enumerate(old_children[remaining_start:], remaining_start):
            self._raise_if_cancelled(cancel_requested)
            differences.append(
                Difference(
                    DifferenceType.DELETED,
                    XML_SHEET_NAME,
                    old_child_paths[index],
                    self._element_summary(child),
                )
            )
        for index, child in enumerate(new_children[remaining_start:], remaining_start):
            self._raise_if_cancelled(cancel_requested)
            differences.append(
                Difference(
                    DifferenceType.ADDED,
                    XML_SHEET_NAME,
                    new_child_paths[index],
                    None,
                    self._element_summary(child),
                )
            )

    @classmethod
    def _child_lcs_matches(
        cls,
        old_children: list[ElementTree.Element],
        new_children: list[ElementTree.Element],
        options: CompareOptions,
        cancel_requested: CancelCheck | None,
    ) -> list[tuple[int, int]]:
        old_signatures = [cls._element_signature(child, options) for child in old_children]
        new_signatures = [cls._element_signature(child, options) for child in new_children]
        row_count = len(old_signatures)
        column_count = len(new_signatures)
        lengths = [[0] * (column_count + 1) for _ in range(row_count + 1)]

        for old_index, old_signature in enumerate(old_signatures, 1):
            cls._raise_if_cancelled(cancel_requested)
            for new_index, new_signature in enumerate(new_signatures, 1):
                if old_signature == new_signature:
                    lengths[old_index][new_index] = lengths[old_index - 1][new_index - 1] + 1
                else:
                    lengths[old_index][new_index] = max(
                        lengths[old_index - 1][new_index],
                        lengths[old_index][new_index - 1],
                    )

        matches: list[tuple[int, int]] = []
        old_index = row_count
        new_index = column_count
        while old_index > 0 and new_index > 0:
            cls._raise_if_cancelled(cancel_requested)
            if old_signatures[old_index - 1] == new_signatures[new_index - 1]:
                matches.append((old_index - 1, new_index - 1))
                old_index -= 1
                new_index -= 1
            elif lengths[old_index - 1][new_index] >= lengths[old_index][new_index - 1]:
                old_index -= 1
            else:
                new_index -= 1
        matches.reverse()
        return matches

    @classmethod
    def _element_signature(cls, element: ElementTree.Element, options: CompareOptions) -> str:
        attributes = (
            sorted(element.attrib.items()) if options.ignore_xml_attribute_order else list(element.attrib.items())
        )
        children = [cls._element_signature(child, options) for child in list(element)]
        return repr(
            (
                cls._display_tag(element.tag),
                attributes,
                cls._normalize_value(cls._text_value(element.text, options), options),
                children,
            )
        )

    @staticmethod
    def _same_tag_sequence(old_children: list[ElementTree.Element], new_children: list[ElementTree.Element]) -> bool:
        return len(old_children) == len(new_children) and all(
            old_child.tag == new_child.tag for old_child, new_child in zip(old_children, new_children, strict=True)
        )

    @classmethod
    def _child_paths(cls, parent_path: str, children: list[ElementTree.Element]) -> list[str]:
        counts: dict[str, int] = {}
        paths: list[str] = []
        for child in children:
            tag = cls._display_tag(child.tag)
            counts[tag] = counts.get(tag, 0) + 1
            paths.append(f"{parent_path}/{tag}[{counts[tag]}]")
        return paths

    @classmethod
    def _element_summary(cls, element: ElementTree.Element) -> str:
        return ElementTree.tostring(element, encoding="unicode", short_empty_elements=True)

    @staticmethod
    def _display_tag(tag: str) -> str:
        if tag.startswith("{") and "}" in tag:
            namespace, local_name = tag[1:].split("}", 1)
            return f"{local_name}{{{namespace}}}"
        return tag

    @classmethod
    def _text_value(cls, value: str | None, options: CompareOptions) -> str | None:
        if value is None:
            return None
        if options.ignore_xml_blank_text and value.strip() == "":
            return None
        return value

    @staticmethod
    def _equal(left: Any, right: Any, options: CompareOptions) -> bool:
        return XmlComparer._normalize_value(left, options) == XmlComparer._normalize_value(right, options)

    @staticmethod
    def _normalize_value(value: Any, options: CompareOptions) -> Any:
        if options.empty_string_equals_empty and value == "":
            value = None
        if isinstance(value, str):
            if options.ignore_surrounding_whitespace:
                value = value.strip()
            if options.ignore_case:
                value = value.casefold()
        return value

    @staticmethod
    def _raise_if_cancelled(cancel_requested: CancelCheck | None) -> None:
        if cancel_requested is not None and cancel_requested():
            raise OperationCancelledError("比較をキャンセルしました。")


class XmlReportWriter(ExcelReportWriter):
    def write(
        self,
        source_new: Path,
        output: Path,
        result: CompareResult,
        detailed: bool = True,
        cancel_requested: CancelCheck | None = None,
        options: CompareOptions | None = None,
    ) -> Path:
        workbook = None
        temporary_output: Path | None = None
        try:
            self._raise_if_cancelled(cancel_requested)
            output.parent.mkdir(parents=True, exist_ok=True)
            file_descriptor, temporary_name = tempfile.mkstemp(
                dir=output.parent,
                prefix=f".{output.stem}_",
                suffix=".tmp.xlsx",
            )
            os.close(file_descriptor)
            temporary_output = Path(temporary_name)

            xml_document = XmlReader().read(source_new)
            workbook = Workbook()
            xml_sheet = workbook.active
            xml_sheet.title = XML_SHEET_NAME
            self._write_xml_sheet(xml_sheet, xml_document, cancel_requested)
            report = workbook.create_sheet(self._unique_report_name(workbook.sheetnames), 0)
            self._write_report(report, self._displayable_result(result), detailed, cancel_requested)
            self._remove_xml_links(report, result)
            self._write_xml_settings(report, options or CompareOptions())
            self._raise_if_cancelled(cancel_requested)
            workbook.save(temporary_output)
            workbook.close()
            workbook = None
            os.replace(temporary_output, output)
            temporary_output = None
            result.output_path = output
            return output
        except (PermissionError, OSError, InvalidFileException, ValueError, TypeError) as exc:
            raise OutputWriteError(f"出力ファイルを保存できません: {output}") from exc
        finally:
            if workbook is not None:
                workbook.close()
            if temporary_output is not None:
                with suppress(OSError):
                    temporary_output.unlink(missing_ok=True)

    def _write_xml_sheet(
        self,
        sheet: Worksheet,
        document: XmlDocument,
        cancel_requested: CancelCheck | None = None,
    ) -> None:
        sheet["A1"] = "新XML"
        sheet["A1"].font = Font(bold=True, size=14)
        text = self._pretty_xml(document.root)
        for row_index, line in enumerate(text.splitlines(), 2):
            self._raise_if_cancelled(cancel_requested)
            self._write_cell(sheet, row_index, 1, line)
        sheet.column_dimensions["A"].width = 120

    @staticmethod
    def _pretty_xml(root: ElementTree.Element) -> str:
        copied = ElementTree.fromstring(ElementTree.tostring(root, encoding="unicode"))
        ElementTree.indent(copied, space="  ")
        return ElementTree.tostring(copied, encoding="unicode", short_empty_elements=True)

    @staticmethod
    def _displayable_result(result: CompareResult) -> CompareResult:
        return CompareResult(
            [
                Difference(
                    difference.kind,
                    difference.sheet,
                    difference.cell,
                    XmlReportWriter._display_xml_value(difference.old_value),
                    XmlReportWriter._display_xml_value(difference.new_value),
                    difference.value_changed,
                    difference.formula_changed,
                )
                for difference in result.differences
            ],
            result.output_path,
        )

    @staticmethod
    def _display_xml_value(value: object) -> object:
        if isinstance(value, list):
            return ", ".join(str(item) for item in value)
        return value

    @staticmethod
    def _remove_xml_links(sheet: Worksheet, result: CompareResult) -> None:
        for row_index in range(11, 11 + result.total):
            cell = sheet.cell(row_index, 6)
            cell.value = None
            cell.hyperlink = None
            cell.style = "Normal"

    @classmethod
    def _write_xml_settings(cls, sheet: Worksheet, options: CompareOptions) -> None:
        sheet["H1"] = "XML読み込み設定"
        sheet["H1"].font = Font(bold=True, size=14)
        rows = [
            ("文字コード", "UTF-8 / UTF-8 BOM"),
            ("比較位置", "XPath風パス"),
            ("属性順を無視", "はい" if options.ignore_xml_attribute_order else "いいえ"),
            ("空白のみテキストを無視", "はい" if options.ignore_xml_blank_text else "いいえ"),
        ]
        for row_index, (label, value) in enumerate(rows, 2):
            cls._write_cell(sheet, row_index, 8, label)
            cls._write_cell(sheet, row_index, 9, value)
        sheet.column_dimensions["H"].width = 22
        sheet.column_dimensions["I"].width = 24
